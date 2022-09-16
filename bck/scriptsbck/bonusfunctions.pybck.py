import csv
import os
import pandas as pd
import json
import logzero
from datetime import datetime
from calendar import isleap
import sys
import os
import xml.etree.ElementTree as ET
import csv
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import xmltodict
from io import BytesIO
import json
from io import BytesIO
from pdfdocument.document import PDFDocument
import json
import csv


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
logs_path = os.path.join(ROOT_DIR, "Logs")
os.makedirs(logs_path, exist_ok=True)
data_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(data_path, exist_ok=True)
logzero.logfile(f"{logs_path}\\functions.log")
logzero.loglevel(logzero.INFO)
logzero.json()

QUARTER1="01/08/2020"
QUARTER1="01/08/2020"
QUARTER1="01/08/2020"
QUARTER1="01/08/2020"

def Create_Workday_Url(year,format):
    try:
        prior_year = int(year)-1
        percent = "%"
        url = f"https://wd5-services1.myworkday.com/ccx/service/customreport2/guidewire/Guidewire_ISU/Bonus_Calculation_Raas?Start_Date={prior_year}-08-01-08{percent}3A00&End_Date={year}-08-01-08{percent}3A00&format={format}"
    except:
        return "Python Error: Error on url <br>"
    else:
        return url
    
def Save_Sheet(Dic, File, Sheet, Col, Row):
    df = pd.DataFrame.from_dict(Dic).T
    try:
        book = load_workbook(File)
    except:
        book = openpyxl.Workbook()
        book.save(File)
        book = load_workbook(File)
    writer = pd.ExcelWriter(File, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=Sheet, startcol=Col, startrow=Row)
    writer.save()
    
    
def Get_Xml_Data(year):
    #try:    
    Xml_Data = open(f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml","r").read()  
    dictionary = xmltodict.parse(Xml_Data)
    json_data = json.dumps(dictionary)
    #print(json_data)
    data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
    #print(data)
    employee_data={}
    count=0
    for i in (data):
            employee_data[count]={
            "Employee_ID":i["wd:Row_Sequence"],
            "Employee_Status":i["wd:Active_Status"],
            "Employee_Type":i["wd:Employee_Type"],
            "Employee_HE_Date":i["wd:Hire_Date"],
            "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
            "Employee_Current_Salary":i["wd:Salary"],
            "Employee_Current_Currency":i["wd:Salary_Currency"],
            "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
            "Employee_Current_Bonus_Amount":i["wd:Bonus_Amount"],
            "Employee_Current_Bonus_Percent":i["wd:Bonus_Percent"],
            "Employee_Current_Commission":i["wd:Commission_Amount"],
            "Employee_Current_Company":i["wd:Company"],
            "Employee_Current_Company_ID":i["wd:Company_ID"],
            "Employee_Current_DPT":i["wd:Cost_Center"],
            "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
            "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
            "Employee_Current_Region":i["wd:Region"],
            "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
            count+=1
            try:
                if type((i)['wd:Compensation_History']) is dict:
                    employee_data[count]={
                    "Employee_ID":i["wd:Row_Sequence"],
                    "Employee_Status":i["wd:Active_Status"],
                    "Employee_Type":i["wd:Employee_Type"],
                    "Employee_HE_Date":i["wd:Compensation_History"]["wd:Effective_Date_Compensation"],
                    "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
                    "Employee_Current_Salary":i["wd:Compensation_History"]["wd:Proposed_Salary"],
                    "Employee_Current_Currency":i["wd:Compensation_History"]["wd:Proposed_Salary_Currency"],
                    "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                    "Employee_Current_Bonus_Amount":i["wd:Compensation_History"]["wd:Proposed_Bonus_Amount"],
                    "Employee_Current_Bonus_Percent":i["wd:Compensation_History"]["wd:Proposed_Bonus_Percent"],
                    "Employee_Current_Commission":i["wd:Commission_Amount"],
                    "Employee_Current_Company":i["wd:Company"],
                    "Employee_Current_Company_ID":i["wd:Company_ID"],
                    "Employee_Current_DPT":i["wd:Cost_Center"],
                    "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                    "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                    "Employee_Current_Region":i["wd:Region"],
                    "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                    count+=1
                elif type((i)['wd:Compensation_History']) is list:
                         for j in range(len(i['wd:Compensation_History'])):
                            employee_data[count]={
                            "Employee_ID":i["wd:Row_Sequence"],
                            "Employee_Status":i["wd:Active_Status"],
                            "Employee_Type":i["wd:Employee_Type"],
                            "Employee_HE_Date":i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"],
                            "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
                            "Employee_Current_Salary":i["wd:Compensation_History"][j]["wd:Proposed_Salary"],
                            "Employee_Current_Currency":i["wd:Compensation_History"][j]["wd:Proposed_Salary_Currency"],
                            "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                            "Employee_Current_Bonus_Amount":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Amount"],
                            "Employee_Current_Bonus_Percent":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Percent"],
                            "Employee_Current_Commission":i["wd:Commission_Amount"],
                            "Employee_Current_Company":i["wd:Company"],
                            "Employee_Current_Company_ID":i["wd:Company_ID"],
                            "Employee_Current_DPT":i["wd:Cost_Center"],
                            "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                            "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                            "Employee_Current_Region":i["wd:Region"],
                            "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                            count+=1
            except :
                pass
            finally:
                print(employee_data)
    Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS.xlsx",f"Bonus", 0, 0)
    return employee_data
                 
        #except: 
        #    pass
            
    
    
      
    
#def Get_Dic(csv_data):
#    dic = {}
#    Flag = 0
#    Count = 0
#    lines = csv_data.splitlines()
#    for l in csv.reader(lines, quotechar='"', delimiter=',', quoting=csv.QUOTE_ALL, skipinitialspace=True):
#        if Flag == 0:
#            Header = l
#            Flag = 1
#        else:
#            dic[Count] = {}
#            for i in range(len(l)):
#                dic[Count][Header[i]] = l[i]
#            Count += 1
#    return dic

def remove_element_by_date(date_of_hire_rehire,cutoffdate):
    dateHR =  datetime.strptime(date_of_hire_rehire, '%m/%d/%y')
    COdate = datetime.strptime(cutoffdate, '%m/%d/%y')
    if dateHR > COdate: 
        return False
    else:
        return True

def remove_element_by_inter(type_oh_employee):
    if type_oh_employee.upper() == "REGULAR":
        return True
    else:
        return False
    
def remove_element_by_active(type_oh_employee):
    if type_oh_employee.upper() == 1:
        return True
    else:
        return False
        
def read_xls():
    try:
        df = pd.read_excel (f'{ROOT_DIR}\\Data\\Bonus_Parameters.xlsx')
        return df
    except:
        return "Error: Cannot read the excel file"
        
        
def get_quarters(year):
    Prior_Year = int(year)-1
    QUARTER1=f"08/01/{Prior_Year}"
    QUARTER2=f"11/01/{Prior_Year}"
    QUARTER3=f"02/01/{year}"
    QUARTER4=f"05/01/{year}"
    return [QUARTER1, QUARTER2, QUARTER3, QUARTER4]

def create_taxes(XLS):
    taxes_dic={}
    for i in range(5,len(XLS["Bonus_Parameters"])):
        taxes_dic[XLS.iloc[i,0]]=[XLS.iloc[i,1],XLS.iloc[i,2],XLS.iloc[i,3],XLS.iloc[i,4]]
    return taxes_dic

def get_employees_with_benefits(Employee_dic,cutoffdate):
    employees_with_benefits = {}
    count = 0
    for i in Employee_dic.keys():
        effectivedate= Employee_dic[i]["Effective_Date"]
        if  remove_element_by_inter(Employee_dic[i]["Employee_Type"]) and remove_element_by_date(f"{effectivedate}",f"{cutoffdate}") and remove_element_by_active(Employee_dic[i]["Active_Status"]):
            employees_with_benefits[count]=Employee_dic[i]
            employees_with_benefits[count] = effectivedate= Employee_dic[i]
            count+=1
            
            
def calculate_bonus(Employee_dic):
    pass