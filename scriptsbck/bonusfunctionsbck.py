import csv
import os
import pandas as pd
import json
import logzero
from datetime import datetime
from calendar import isleap
import sys
import os
import xml.etree.ElementTree as ET
import csv
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import xmltodict
from io import BytesIO
from operator import itemgetter
import json
from io import BytesIO
from pdfdocument.document import PDFDocument
import json
import csv
import requests
from dateutil.relativedelta import relativedelta
from logzero import logger
import base64

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
logs_path = os.path.join(ROOT_DIR, "Logs")
os.makedirs(logs_path, exist_ok=True)
data_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(data_path, exist_ok=True)
config_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(config_path, exist_ok=True)
logzero.logfile(f"{logs_path}\\functions.log")
logzero.loglevel(logzero.INFO)
logzero.json()
#oracle_url="https://efow.fs.us2.oraclecloud.com"
#oracle_url="https://efow-test.fa.us2.oraclecloud.com"
oracle_url= "https://efow-dev1.fa.us2.oraclecloud.com"


set_id = "300000001414463"
#Autopost_id = "300000150481649"
#source_id = "300000150481632"
######## DEV DATA
Autopost_id = "300000146748162"
source_id= "300000146748158"

QUARTER1S="01/08/2020"
QUARTER2S="01/08/2020"
QUARTER3S="01/08/2020"
QUARTER4S="01/08/2020"

months = (["Jan", "Feb", "Mar", "Apr",
           "May", "Jun", "Jul", "Aug",
           "Sep", "Oct", "Nov", "Dec"]
          )


   
def set_user(data):
    logger.info("Setting credentials")
    try:
       user,passw= data.split("///")
       with open(f"{ROOT_DIR}\\Data\\credentials.txt","wb") as cred:
            cred.write(base64.b64encode(user.encode()))
            cred.write(b"\n")
            cred.write(base64.b64encode(passw.encode()))
            cred.close()
    except:
        return "Error on set credentials // //"




def read_credentials():
    try:
        usr_psw = open(f"{ROOT_DIR}\\Data\\credentials.txt","rb")
        user = base64.b64decode(usr_psw.readline().decode()).decode()
        passw = base64.b64decode(usr_psw.readline().decode()).decode()
        usr_psw.close()
    except:
        return "Error on read credentials // //"
    return user,passw



def get_ledgers_info():
    try:
        logger.info("Getting ledgers info to look for closed ones")
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/ledgersLOV?limit=10000"
        #print(url)
        headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
        usr_psw = open(f"{ROOT_DIR}\\Data\\credentials.txt","rb")
        user = base64.b64decode(usr_psw.readline().decode()).decode()
        passw = base64.b64decode(usr_psw.readline().decode()).decode()
    except:
        return "Error on read ledger info // //"
    return requests.get(url, auth=(user, passw), headers=headerss)




def get_entities():
    logger.info("Reading Ledgers Id from Oracle")
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        for i in ledger_json:
            ledgers[i["Name"]] = i["LedgerId"]
    except:
        return "Error on set the entities // // "
    return  {
        "10": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "12": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "14": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "15": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "16": {"Name": "GW CA PL", "Ledger":  ledgers["GW CA PL"]},
        "65": {"Name": "GW BR PL", "Ledger":  ledgers["GW BR PL"]},
        "41": {"Name": "GW UK PL", "Ledger":  ledgers["GW UK PL"]},
        "43": {"Name": "GW FR PL", "Ledger":  ledgers["GW FR PL"]},
        "45": {"Name": "GW DE PL", "Ledger":  ledgers["GW DE PL"]},
        "46": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "47": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "48": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "49": {"Name": "GW IT PL", "Ledger":  ledgers["GW IT PL"]},
        "51": {"Name": "GW PL PL", "Ledger":  ledgers["GW PL PL"]},
        "55": {"Name": "GW CH PL", "Ledger":  ledgers["GW CH PL"]},
        "57": {"Name": "GW ES PL", "Ledger":  ledgers["GW ES PL"]},
        "71": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "72": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "80": {"Name": "GW CN PL", "Ledger":  ledgers["GW CN PL"]},
        "86": {"Name": "GW JP PL", "Ledger":  ledgers["GW JP PL"]},
        "84": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "85": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "82": {"Name": "GW MY PL", "Ledger":  ledgers["GW MY PL"]},
        "61": {"Name": "GW AR PL", "Ledger":  ledgers["GW AR PL"]},
        "53": {"Name": "GW AT PL", "Ledger":  ledgers["GW AT PL"]},
        "58": {"Name": "GW DK PL", "Ledger":  ledgers["GW DK PL"]}, 
    }



'''
entities = {
                        "10": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "12": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "14": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "15": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "16": {"Name": "GW CA PL", "Ledger": "300000001414017"},
                        "65": {"Name": "GW BR PL", "Ledger": "300000001414018"},
                        "41": {"Name": "GW UK PL", "Ledger": "300000001414019"},
                        "43": {"Name": "GW FR PL", "Ledger": "300000001414020"},
                        "45": {"Name": "GW DE PL", "Ledger": "300000001414021"},
                        "46": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "47": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "48": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "49": {"Name": "GW IT PL", "Ledger": "300000001414023"},
                        "51": {"Name": "GW PL PL", "Ledger": "300000001414024"},
                        "55": {"Name": "GW CH PL", "Ledger": "300000001414025"},
                        "57": {"Name": "GW ES PL", "Ledger": "300000001414026"},
                        "71": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "72": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "80": {"Name": "GW CN PL", "Ledger": "300000001414029"},
                        "86": {"Name": "GW JP PL", "Ledger": "300000001414030"},
                        "84": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "85": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "82": {"Name": "GW MY PL", "Ledger": "300000019642549"},
                        "61": {"Name": "GW AR PL", "Ledger": "300000019833039"},
                        "53": {"Name": "GW AT PL", "Ledger": "300000017944094"},
                        "58": {"Name": "GW DK PL", "Ledger": "300000026789264"},
                        }
'''

def Create_Workday_Url(Workday_String_Start_date,String_Cutoff,format):
    logger.info("Creating Workday Endpoint")
    try:
        percent = "%"
        url = f"https://wd5-services1.myworkday.com/ccx/service/customreport2/guidewire/Guidewire_ISU/Bonus_Calculation_Raas?Start_Date={Workday_String_Start_date}-08{percent}3A00&End_Date={String_Cutoff}-08{percent}3A00&format={format}"
    except:
        return "Python Error: Error on url <br> // // "
    else:
        return url
    
    
def Save_Sheet(Dic, File, Sheet, Col, Row):
    logger.info("Saving Data in Xlsx")
    try:
        df = pd.DataFrame.from_dict(Dic).T
    except:
        return "Error on get the dictionary // // "
    else:
        try:
            book = load_workbook(File)
        except:
            book = openpyxl.Workbook()
            book.save(File)
            book = load_workbook(File)
        writer = pd.ExcelWriter(File, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=Sheet, startcol=Col, startrow=Row)
        writer.save()
        
            
        
def read_xls_parameters():
    logger.info("Reading Parameters")
    try:
        df = pd.read_excel (f'{ROOT_DIR}\\Data\\Bonus_Parameters.xlsx')
        return df
    except:
        return "Error: Cannot read the excel file // // "

        
        
def get_quarters(Year):
    logger.info("Creating predefined quarters")
    try:
        Prior_Year = int(Year)-1
        QUARTER1S=f"08/01/{Prior_Year}"
        QUARTER2S=f"11/01/{Prior_Year}"
        QUARTER3S=f"02/01/{Year}"
        QUARTER4S=f"05/01/{Year}"
        QUARTER1E=f"10/31/{Prior_Year}"
        QUARTER2E=f"01/31/{Year}"
        QUARTER3E=f"04/30/{Year}"
        QUARTER4E=f"07/31/{Year}"
    except:
        return "Error on set the quarters // // "
    return [(datetime.strptime(QUARTER1S,"%m/%d/%Y"),datetime.strptime(QUARTER1E,"%m/%d/%Y")) ,(datetime.strptime(QUARTER2S,"%m/%d/%Y"),datetime.strptime(QUARTER2E,"%m/%d/%Y")), (datetime.strptime(QUARTER3S,"%m/%d/%Y"),datetime.strptime(QUARTER3E,"%m/%d/%Y")), (datetime.strptime(QUARTER4S,"%m/%d/%Y"),datetime.strptime(QUARTER4E,"%m/%d/%Y"))]

def create_taxes(XLS):
    logger.info("Reading Taxes Info")
    try:
        taxes_dic={}
        for i in range(4,len(XLS["Bonus_Parameters"])):
            taxes_dic[XLS.iloc[i,0]]=[XLS.iloc[i,1],XLS.iloc[i,2],XLS.iloc[i,3],XLS.iloc[i,4]]
    except:
        return "Error on create taxes // // "
    return taxes_dic
            

def get_employee_data_file(url,year,format):
    logger.info("Reading WorkDay data and saving in file")
    try:
        with open(f"{data_path}\\{year}_bonusfile.{format}", "w+b") as xml_to_write:
            bonus_data = requests.get(f"{url}", auth=(
                                                "SVC_RPA_WDAY",
                                                "Gu1d3wire!"
                                                ),
                                            stream=True).content
            xml_to_write.write(bonus_data)
            return "0"
    except:
        return "Error when trying to save the workday data // // "
    
    
    
def Get_Xml_Data(year):
    logger.info("Reading xml file")
    try:    
        Xml_Data = open(f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml","r").read()  
        dictionary = xmltodict.parse(Xml_Data)
        json_data = json.dumps(dictionary)
        data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
        employee_data={}
        count=0
        try:
            logger.info("Extracting Employee data")
            for i in (data):
                    #print(i)
                    employee_data[count]={
                    "Employee_ID":i["wd:Row_Sequence"],
                    "Employee_Status":i["wd:Active_Status"],
                    "Employee_Type":i["wd:Employee_Type"],
                    "Employee_HE_Date":i["wd:Hire_Date"],
                    "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
                    "Employee_Current_Salary":i["wd:Salary"],
                    "Employee_Current_Currency":i["wd:Salary_Currency"],
                    "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                    "Employee_Current_Bonus_Amount":i["wd:Bonus_Amount"],
                    "Employee_Current_Bonus_Percent":i["wd:Bonus_Percent"],
                    "Employee_Current_Commission":i["wd:Commission_Amount"],
                    "Employee_Current_Company":i["wd:Company"],
                    "Employee_Current_Company_ID":i["wd:Company_ID"],
                    "Employee_Current_DPT":i["wd:Cost_Center"],
                    "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                    "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                    "Employee_Current_Region":i["wd:Region"],
                    "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                    count+=1
                    try:
                        if type((i)['wd:Compensation_History']) is dict:
                            employee_data[count]={
                            "Employee_ID":i["wd:Row_Sequence"],
                            "Employee_Status":i["wd:Active_Status"],
                            "Employee_Type":i["wd:Employee_Type"],
                            "Employee_HE_Date":i["wd:Compensation_History"]["wd:Effective_Date_Compensation"],
                            "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
                            "Employee_Current_Salary":i["wd:Compensation_History"]["wd:Proposed_Salary"],
                            "Employee_Current_Currency":i["wd:Compensation_History"]["wd:Proposed_Salary_Currency"],
                            "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                            "Employee_Current_Bonus_Amount":i["wd:Compensation_History"]["wd:Proposed_Bonus_Amount"],
                            "Employee_Current_Bonus_Percent":i["wd:Compensation_History"]["wd:Proposed_Bonus_Percent"],
                            "Employee_Current_Commission":i["wd:Commission_Amount"],
                            "Employee_Current_Company":i["wd:Company"],
                            "Employee_Current_Company_ID":i["wd:Company_ID"],
                            "Employee_Current_DPT":i["wd:Cost_Center"],
                            "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                            "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                            "Employee_Current_Region":i["wd:Region"],
                            "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                            count+=1
                        elif type((i)['wd:Compensation_History']) is list:
                                for j in range(len(i['wd:Compensation_History'])):
                                    employee_data[count]={
                                    "Employee_ID":i["wd:Row_Sequence"],
                                    "Employee_Status":i["wd:Active_Status"],
                                    "Employee_Type":i["wd:Employee_Type"],
                                    "Employee_HE_Date":i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"],
                                    "Employee_PF_Time":i["wd:Time_Type"]['@wd:Descriptor'],
                                    "Employee_Current_Salary":i["wd:Compensation_History"][j]["wd:Proposed_Salary"],
                                    "Employee_Current_Currency":i["wd:Compensation_History"][j]["wd:Proposed_Salary_Currency"],
                                    "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                                    "Employee_Current_Bonus_Amount":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Amount"],
                                    "Employee_Current_Bonus_Percent":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Percent"],
                                    "Employee_Current_Commission":i["wd:Commission_Amount"],
                                    "Employee_Current_Company":i["wd:Company"],
                                    "Employee_Current_Company_ID":i["wd:Company_ID"],
                                    "Employee_Current_DPT":i["wd:Cost_Center"],
                                    "Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                                    "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                                    "Employee_Current_Region":i["wd:Region"],
                                    "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                                    count+=1
                    except :
                        pass
            #print(employee_data)
            #Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS.xlsx",f"Data_From_WorkDay", 0, 0)
            #return employee_data
        except:
            pass
             #return "Error at time to read xml data //"
    except: 
         return "Error at time to read the XML file // // "
    else:
        Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Data_From_WorkDay", 0, 0)
        return employee_data
        
    

def get_dict_with_condition(employee_data,field,value):
    logger.info("Removing data that doesn't complies with condition")
    try:
        employees_with_benefits = {}
        count = 0
        for i in employee_data.keys():
            if str(employee_data[i][field]).upper() == str(value).upper():
                employees_with_benefits[count]=employee_data[i]
                count+=1
    except:
        return "Error : Cant set the dictionary // // "
    return employees_with_benefits




   
def get_data_before_cutoff(employee_data,cutoffdate):
    logger.info("removing data before cutoff")
    try:
        employees_with_benefits = {}
        count = 0
        for i in employee_data.keys():
            #print(employee_data[i]["Employee_HE_Date"][0:-6],cutoffdate)
            dateHR =  datetime.strptime(employee_data[i]["Employee_HE_Date"][0:-6], '%Y-%m-%d')
            if dateHR <= cutoffdate:
                employees_with_benefits[count]=employee_data[i]
                count+=1
    except:
        return "Error cant get employees before cutoff // // "
    return employees_with_benefits


def order_dic(dic):
    logger.info("Ordering data")
    try:
        list_from_dict = sorted(dic.values(), key=itemgetter("Employee_ID", "Employee_HE_Date"), reverse=True)
        out_dic = {}
        for i in range(len(list_from_dict)):
            out_dic[i] = list_from_dict[i]
    except:
        return "Error : cant order dictionary // // "
    return out_dic
   
   
     
    

def create_pivot_data(Employee_dic,quarters,proration,taxes,cutoff,quarter_to_calculate,company_performances):
    logger.info("Creating pre-pivot")
    #try:
    currencies = []
    first_id = -1
    Original_Journal = {}
    Converted_Journal = {}
    data = []
    first_data =[]
    count = 0
    last_date = None
    for i in Employee_dic.keys():
        first_data = get_employee_data_current_employee(Employee_dic[i])
        employee_current_id = int(first_data[0])
        if  employee_current_id != first_id:
            data = first_data
            first_id = employee_current_id 
        #if first_data[4] not in currencies:
        #    currencies.append(first_data[4])
        temporal_item = get_partial_amount(quarter_to_calculate, first_data, data, quarters,proration,cutoff,company_performances)
        if temporal_item != {0}:
            Original_Journal[count] = temporal_item 
            count +=1
    #except:
    #    return "Error Cant create the pivot data // // "
    return Original_Journal
    
   
def get_employee_data_current_employee(Employee):
    return [Employee['Employee_ID'], Employee["Employee_HE_Date"],  Employee["Employee_Current_Company_ID"],  Employee["Employee_Current_DPT"], Employee["Employee_Current_Currency"], Employee["Employee_Current_Bonus_Plan"], Employee["Employee_Current_Salary"], Employee["Employee_Current_Bonus_Percent"], Employee[ "Employee_Current_Commission"]]

def get_partial_amount(quarter_to_calculate, first_data, data, quarters,proration,cutoff,company_performances):
    #try:
    String_date_Formatter = "%Y-%m-%d"
    #print(data, "\n--->\n", first_data,"\n\n\n\n")
    if  "Fixed Percent" in first_data[5]:
        if data == first_data:
            dpt= int(first_data[3])
            Entity = int(first_data[2])
            currency = first_data[4]
            bonus_amount= float(first_data[6]) * float(first_data[7])
            #create journal directly
        else:
            dpt = int(data[3])
            Entity = int(data[2])
            from_currency= first_data[4]
            currency = data[4]
            end_date = quarters[quarter_to_calculate-1][1]
            #print(from_currency,currency,end_date)
            convertion_rate = get_currency_exchange(from_currency,currency,end_date.strftime(String_date_Formatter))
            print(f"conversion: {convertion_rate}")
            bonus_amount= float(first_data[6]) * float(first_data[7])
            bonus_amount = float(bonus_amount) * float(convertion_rate)
    else:
        if data == first_data:
            dpt= int(first_data[3])
            Entity = int(first_data[2])
            currency = first_data[4]
            bonus_amount= float(first_data[8])
        else:
            dpt = int(data[3])
            Entity = int(data[2])
            from_currency= first_data[4]
            currency = data[4]
            #print(from_currency,currency)
            convertion_rate = get_currency_exchange(from_currency,currency,cutoff.strftime(String_date_Formatter))
            print(f"conversion: {convertion_rate}")
            bonus_amount= float(first_data[8]) * float(convertion_rate)
    #print("Ammount_from_Workday",bonus_amount)
    total_amount = 0
    for i in range(quarter_to_calculate-1,-1,-1):
        HR_date=  datetime.strptime(first_data[1],"%Y-%m-%d-%H:%M")
        percen_quarter_worked = calculate_quarter_proration(quarters[i][0],HR_date ,quarters[i][1])
        #print("Percent_of Quarter",percen_quarter_worked)
        partial_calculation =  (float(bonus_amount) * company_performances) / 365 
        #print("Partial Calculation",partial_calculation)
        #print("Proration", float(proration[i]))
        #print("Subtotal",partial_calculation * percen_quarter_worked * float(proration[i]))
        total_amount += partial_calculation * percen_quarter_worked * float(proration[i])
        #print("Partial Total Ammount",total_amount)
    total_amount = round(total_amount)
    #print("Rounded",total_amount)
    #print({'Entity' : Entity ,'Dpt': dpt, 'Currency' : currency , 'Amount' : total_amount })
    if total_amount > 0:
        return  {'Entity' : Entity ,'Dpt': dpt, 'Currency' : currency , 'Amount' : total_amount }  
    else:
        return {0}
    #except:
    #    return "Error Cant get partial amount // // "
        
        
def calculate_quarter_proration(start_date,date_of_hire_rehire,end_date):
    try:
        if date_of_hire_rehire > end_date:
            return 0
        elif date_of_hire_rehire < start_date:
            return 1
        else:
            return float ( (date_of_hire_rehire - start_date).days / (end_date - start_date).days )
    except:
        return "Error cant calculate quarter proration // // "
    
def get_currency_exchange(from_currency,to_currency,date_of_convertion):
    #try:
    if from_currency == to_currency:
        return 1
    currentDateTime = datetime.now() - relativedelta(months=1) 
    if type(date_of_convertion) == str: 
        date_of_convertion = datetime.strptime(date_of_convertion, '%Y-%m-%d')
    if date_of_convertion > currentDateTime:
        date_of_convertion = currentDateTime.strftime('%Y-%m-%d')
    else:
        date_of_convertion = date_of_convertion.strftime('%Y-%m-%d')
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/currencyRates?finder=CurrencyRatesFinder;fromCurrency={from_currency},toCurrency={to_currency},userConversionType=Daily Rate,startDate={date_of_convertion},endDate={date_of_convertion}&onlyData=True"
    headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
    user,passw  = read_credentials()
    print(user,passw)
    print(url)
    user = "vsethi"
    passw = "Welcome1"
    ret = requests.get(url, auth=(user, passw), headers=headerss)
    print(ret)
    print(ret.text)
    #print(from_currency,to_currency,date_of_convertion)
    rate = float(json.loads(ret.text)["items"][0]["ConversionRate"])
    print(rate)
    #except:
    #    return "Error : cant calculate rate convertion // // "
    return rate


def create_pivot(pivot_data):
    logger.info("Creating Pivot")
    pivot={}
    #try:
    for i in pivot_data.keys():
        print(f"number : {i}")
        print(f"data: {pivot_data[i]}")
        print(f'entity: {pivot_data[i]["Entity"]}')
        entity = int(pivot_data[i]["Entity"])
        dpt= int(pivot_data[i]["Dpt"])
        currency = pivot_data[i]["Currency"]
        amount= round(float(pivot_data[i]["Amount"]), 2)
        
        if entity not in pivot.keys():
            pivot[entity] =  {  dpt: { 'Currency': currency, 'Amount' : round(float(amount),2) } }
        else:
            if dpt in pivot[entity].keys():
                total_amount = round(float(pivot[entity][dpt]["Amount"]),2) + amount
                pivot[entity][dpt]= { 'Currency': currency, 'Amount' : round(float(total_amount),2)}  
            else:
                pivot[entity].update({ dpt : { 'Currency': currency, 'Amount' : round(amount,2) } })  
    #except:
    #      return "Error can create the pivot // // "    
    return pivot


        
        
        


def create_Journals(pivot,end_date,reversal,quarter_to_calculate,taxes):
    logger.info("creating Journal")
    #try:
    entities=get_entities()
    jrnl={}
    tx_jrn={}
    usd_jrnl={}
    tx_usd_jrnl={}
    count=0
    Year_To_calculate = int(end_date.strftime("%Y"))
    year = f"{Year_To_calculate}"[:-2]
    Base_Bach_Name= f"Q{quarter_to_calculate}'{Year_To_calculate} - Bonus Accrual"
    Batch_Name = Base_Bach_Name
    Batch_Description = Base_Bach_Name
    #print(pivot) 
    period = end_date.strftime("%b-%y")
    state= Get_Ledger_Status(f"{period}")
    if "Error" in state:
        return f"{state} // // "
    for i in pivot.keys():
        #print(i)
        data = entities[f"{i}"]
        group_id= datetime.today().strftime('%Y%m%d%H%M%S')
        Ledger_Country = data["Name"].split(" ")[1]
        Journal_Name = f"{Base_Bach_Name} - {Ledger_Country}"
        Ledger_Description =  f"{Base_Bach_Name} - {Ledger_Country} {group_id}"
        Ledger_tax_Description =  f"{Base_Bach_Name} - Tax - {Ledger_Country} {group_id}"
        Ledger_converted_Description =  f"{Base_Bach_Name} - Adjustment - {Ledger_Country} {group_id}"
        Ledger_tax_converted_Description =  f"{Base_Bach_Name} - Tax Adjustment -{Ledger_Country} {group_id}"
        Line_Description =  f"{Base_Bach_Name} - {Ledger_Country}"
        Ledger_name = data["Name"]
        Ledger_id=data["Ledger"]
        Accounting_date = end_date.strftime("%Y-%m-%d")
        #datetime.datetime.now().strftime("%m/%d/%Y")
        Source = "AA Bot"
        Category  = "COMPandBEN_Accrual"
        entity = i
        #Line_Description = entities[f"first_data[2]"]
        entity_total=0.00
        total_tax=0.00
        converted_total=0.00
        tax_converted_total=0.00
        header_soap = Create_SOAP_HEADER(Batch_Name,Ledger_id,period,Accounting_date)
        header_soap_tx = Create_SOAP_HEADER(f"{Batch_Name}_Tax",Ledger_id,period,Accounting_date)
        header_soap_us = Create_SOAP_HEADER(f"{Batch_Name}-Adjustment",Ledger_id,period,Accounting_date)
        header_soap_us_tx = Create_SOAP_HEADER(f"{Batch_Name} Tax-Adjustment",Ledger_id,period,Accounting_date)
        journal_line=""
        journal_tax_line=""
        journal_converted_line =""
        journal_tax_converted_tax_line =""
        for j in pivot[i].keys():
            dpt=j
            account="60200"
            tx_account="60810"
            prod="000000"
            proj="000000"
            ico="00"
            fut="0"
            currency= pivot[i][j]["Currency"]
            debit= float(pivot[i][j]["Amount"])
            debit = f"{debit:.2f}"
            debit = float(debit)
            tax =  debit * float(taxes[i][quarter_to_calculate-1])
            tax = f"{tax:.2f}"
            tax = float(tax)    
            convertion_rate = get_currency_exchange(currency,"USD",end_date)
            converted = debit * convertion_rate
            converted = f"{converted:.2f}"
            converted = float(converted)
            tax_converted = tax * convertion_rate
            tax_converted = f"{tax_converted:.2f}"
            tax_converted = float(tax_converted)
            print(f"Entity Total Before: {entity_total}, Total Tax Before: {total_tax}, Converted Total before: {converted_total} , Tx Converted Total Before : {tax_converted_total}")
            entity_total += debit
            entity_total = f"{entity_total:.2f}"
            entity_total = float(entity_total)
            total_tax +=  tax
            total_tax = f"{total_tax:.2f}"
            total_tax = float(total_tax)
            converted_total +=  converted
            converted_total = f"{converted_total:.2f}"
            converted_total = float(converted_total)
            tax_converted_total += tax_converted
            tax_converted_total = f"{tax_converted_total:.2f}"
            tax_converted_total = float(tax_converted_total) 
            print(f"Entity:{entity}, Dpt: {dpt}, Debit: {debit}, Tx: {tax}, Convertion: {converted}, Tax Converted:{tax_converted}, Conversion:{convertion_rate}")
            print(f"Entity Total: {entity_total}, Total Tax: {tax}, Converted Total: {converted_total} , Tx Converted Total: {tax_converted_total}")
            
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, debit, reversal, account, entity, dpt, Journal_Name,Ledger_name,Ledger_Description,0)
            journal_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, tax, reversal, tx_account, entity, dpt, f"{Journal_Name}_Tax",Ledger_name,Ledger_tax_Description,0)
            journal_converted_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", converted, reversal, account, entity, dpt, f"{Journal_Name} - Adjustment",Ledger_name,Ledger_converted_Description,0)
            journal_tax_converted_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", tax_converted, reversal, tx_account, entity, dpt, f"{Journal_Name} Tax- Adjustment",Ledger_name,Ledger_tax_converted_Description,0)
            
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : debit,"Credit":"0", "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description }
            tx_jrn[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : tax,"Credit":"0", "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name}_Tax", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_Description}
            usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : converted,"Credit":"0", "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_converted_Description}
            tx_usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : tax_converted,"Credit":"0", "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} Tax - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_converted_Description}
            count += 1
            
        dpt = "000"
        account="21030"
        tx_account="21110"
        print(f"entity:{entity}, Dpt:{dpt}, Entity Total: {entity_total}, Total Taxes: {total_tax}, Total Converted : {converted_total}, Total Tax Converted: {tax_converted_total}, Conversion:{convertion_rate}")
        journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, entity_total, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
        journal_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, total_tax, reversal, tx_account, entity, dpt, f"{Journal_Name}_Tax",Ledger_name,Ledger_tax_Description,1)
        journal_converted_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", converted_total, reversal, account, entity, dpt, f"{Journal_Name} - Adjustment",Ledger_name,Ledger_converted_Description,1)
        journal_tax_converted_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", tax_converted_total, reversal, tx_account, entity, dpt, f"{Journal_Name} Tax- Adjustment",Ledger_name,Ledger_tax_converted_Description,1)
        
        jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":entity_total, "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Ledger_Name" : Ledger_name,"Journal_description": Ledger_Description}
        tx_jrn[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":total_tax, "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name}_Tax", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_Description}
        usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":converted_total, "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_converted_Description}
        tx_usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":tax_converted_total, "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} Tax - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_converted_Description}
        count += 1
        
        foot= Create_SOAP_Foot()
        journal_entry = f"{header_soap}{journal_line}{foot}"
        journal_tx_entry = f"{header_soap_tx}{journal_tax_line}{foot}"
        journal_converted_entry = f"{header_soap_us}{journal_converted_line}{foot}"
        journal_tx_converted_entry = f"{header_soap_us_tx}{journal_tax_converted_tax_line}{foot}"
        logger.info("Journal lines pushed")
        response_to_push = Push_To_Oracle(journal_entry)
        if ">0</result>" not in response_to_push.text:
            print(response_to_push.text)
            print("Error on pushing")
            return "Error on pushed Journal lines // // "
        logger.info("importing journal lines")
        response_to_push = Push_To_Oracle(journal_tx_entry)
        if ">0</result>" not in response_to_push.text:
            print(response_to_push.text)
            print("Error on pushing")
            return "Error on pushed Journal lines // // "
        response_to_push = Push_To_Oracle(journal_converted_entry)
        if ">0</result>" not in response_to_push.text:
            print(response_to_push.text)
            print("Error on pushing")
            return "Error on pushed Journal lines // // "
        response_to_push = Push_To_Oracle(journal_tx_converted_entry)
        if ">0</result>" not in response_to_push.text:
            print(response_to_push.text)
            print("Error on pushing")
            return "Error on pushed Journal lines // // "
        importing = Create_Import_SOAP(set_id, source_id,Ledger_id , f"{group_id}")
        response_import = Import_to_GL(importing)
        import_id = json.loads(response_import.text)["ReqstId"]
        print(import_id)
        if import_id != -1:
            while True:
                Status = json.loads(Get_Status(import_id).text)["items"][0]["RequestStatus"]
                print(Status)
                if Status == "SUCCEEDED":
                    print("Imported")
                    logger.info("journal lines imported")
                    break
                if Status == "ERROR" or Status == "WARNING":
                    print("Error on import")
                    return f"Python Error: Warning or Error on Importing PID : {import_id}   Ledger ID: {Ledger_id} on journal // //"
                    #return "Python Error: Warning or Error on Importing"
        else:
            logger.info("Error on import Journal lines")
            print("Oracle's Error on import")
            return "Error on import Journal lines // //"
        with open(f"{data_path}\\journal.txt","a+") as jrn:
            jrn.write(journal_entry)
    return jrnl,tx_jrn,usd_jrnl,tx_usd_jrnl
    #except:
    #    return "Error cant create journal // //"





'''
def create_Taxes_Journals(pivot,end_date,reversal,quarter_to_calculate,taxes):
    try:
        entities=get_entities()
        jrnl={}
        count=0
        Year_To_calculate = int(end_date.strftime("%Y")) -1 
        year = f"{Year_To_calculate}"[:-2]
        Base_Bach_Name= f"Q{quarter_to_calculate}'{Year_To_calculate} - Bonus Accrual"
        Batch_Name = Base_Bach_Name
        Batch_Description = Base_Bach_Name
        #print(pivot) 
        period = end_date.strftime("%b-%y")
        for i in pivot.keys():
            data = entities[f"{i}"]
            group_id= datetime.today().strftime('%Y%m%d%H%M%S')
            Ledger_Country = data["Name"].split(" ")[1]
            Journal_Name = f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_Description =  f"{Base_Bach_Name} - {Ledger_Country} {group_id}"
            Line_Description =  f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_name = data["Name"]
            Ledger_id=data["Ledger"]
            Accounting_date = end_date.strftime("%Y-%m-%d")
            #datetime.datetime.now().strftime("%m/%d/%Y")
            Source = "AA Bot"
            Category  = "COMPandBEN_Accrual"
            entity = i
            #Line_Description = entities[f"first_data[2]"]
            entity_total=0
            header_soap = Create_SOAP_HEADER(Batch_Name,Ledger_id,period,Accounting_date)
            journal_line=""
            for j in pivot[i].keys():
                #print(i,)
                dpt=j
                account="60810"
                prod="000000"
                proj="000000"
                ico="00"
                fut="0"
                currency= pivot[i][j]["Currency"]
                debit= round(float(pivot[i][j]["Amount"]) * float(taxes[i][quarter_to_calculate-1]),2)
                print(debit)
                entity_total = entity_total+debit
                #print(currency,amount,entity_total)
                journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, debit, reversal, account, entity, dpt, Journal_Name,Ledger_name,Ledger_Description,0)
                jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : debit,"Credit":"0", "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description }
                count += 1
            dpt = "000"
            account="21110"
            credit= entity_total
            print(credit)
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, credit, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":credit, "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description": Ledger_Description}
            count += 1
            foot= Create_SOAP_Foot()
            journal_entry = f"{header_soap}{journal_line}{foot}"
            response_to_push = Push_To_Oracle(journal_entry)
            if ">0</result>" not in response_to_push.text:
                print("SOAP Error")
                print(response_to_push.text)
                return "Error  on soapJournal Taxes // // "
            importing = Create_Import_SOAP(set_id, source_id,Ledger_id , f"{group_id}")
            response_import = Import_to_GL(importing)
            import_id = json.loads(response_import.text)["ReqstId"]
            print(import_id)
            if import_id != -1:
                while True:
                    Status = json.loads(Get_Status(import_id).text)["items"][0]["RequestStatus"]
                    print(Status)
                    if Status == "SUCCEEDED":
                        print("Imported")
                        break
                    if Status == "ERROR" or Status == "WARNING":
                        print("Error on import")
                        return f"Python Error: Warning or Error on Importing  PID : {import_id}  Ledger ID: {Ledger_id}  on Taxes // //"
                        #return "Python Error: Warning or Error on Importing"
            else:
                print("Oracle's Error on import")
                return "Error on import Journal Taxes// //"
            
            with open(f"{data_path}\\journal_taxes.txt","a+") as jrn:
                jrn.write(journal_entry)
        return jrnl
    except:
        return "Error cant create Taxes journal // //"



def create_Dollar_Journals(pivot,end_date,reversal,quarter_to_calculate):
    try:
        entities=get_entities()
        jrnl={}
        count=0
        Year_To_calculate = int(end_date.strftime("%Y"))
        year = f"{Year_To_calculate}"[:-2]
        Base_Bach_Name= f"Q{quarter_to_calculate}'{Year_To_calculate} - Bonus Accrual"
        Batch_Name = Base_Bach_Name
        Batch_Description = Base_Bach_Name
        #print(pivot) 
        currency="USD"
        period = end_date.strftime("%b-%y")
        for i in pivot.keys():
            print(i)
            data = entities[f"{i}"]
            group_id= datetime.today().strftime('%Y%m%d%H%M%S')
            Ledger_Country = data["Name"].split(" ")[1]
            Journal_Name = f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_Description =  f"{Base_Bach_Name} - {Ledger_Country} {group_id}"
            Line_Description =  f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_name = data["Name"]
            Ledger_id=data["Ledger"]
            Accounting_date = end_date.strftime("%Y-%m-%d")
            #datetime.datetime.now().strftime("%m/%d/%Y")
            Source = "AA Bot"
            Category  = "COMPandBEN_Accrual"
            entity = i
            #Line_Description = entities[f"first_data[2]"]
            entity_total=0
            header_soap = Create_SOAP_HEADER(Batch_Name,Ledger_id,period,Accounting_date)
            journal_line=""
            for j in pivot[i].keys():
                dpt=j
                account="60200"
                prod="000000"
                proj="000000"
                ico="00"
                fut="0"
                from_currency= pivot[i][j]["Currency"]
                convertion_rate = get_currency_exchange(from_currency,currency,end_date)
                debit= round(float(pivot[i][j]["Amount"]) * float(convertion_rate),2)
                print(debit)
                entity_total = entity_total+debit
                #print(currency,amount,entity_total)
                journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, debit, reversal, account, entity, dpt, Journal_Name,Ledger_name,Ledger_Description,0)
                jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Acounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : debit,"Credit":"0", "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description}
                count += 1
            dpt = "000"
            account="21030"
            credit=entity_total
            print(credit)
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, credit, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Acounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":credit, "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description}
            count += 1
            foot= Create_SOAP_Foot()
            journal_entry = f"{header_soap}{journal_line}{foot}"
            response_to_push = Push_To_Oracle(journal_entry)
            if ">0</result>" not in response_to_push.text:
                print("Error on soap")
                print(response_to_push.text)
                return "Error  on soap Journal USD// //"
            importing = Create_Import_SOAP(set_id, source_id,Ledger_id , f"{group_id}")
            response_import = Import_to_GL(importing)
            import_id = json.loads(response_import.text)["ReqstId"]
            print(import_id)
            if import_id != -1:
                while True:
                    Status = json.loads(Get_Status(import_id).text)["items"][0]["RequestStatus"]
                    print(Status)
                    if Status == "SUCCEEDED":
                        print("Imported")
                        break
                    if Status == "ERROR" or Status == "WARNING":
                        print("Error on import")
                        return f"Python Error: Warning or Error on Importing PID : {import_id}    Ledger ID: {Ledger_id}   on converted // //"
            else:
                print("Oracle's Error on import")
                return "Error on importing Journal USD// //"
            
            with open(f"{data_path}\\Dollar_journal.txt","a+") as jrn:
                jrn.write(journal_entry)
        return jrnl
    except:
        return "Error can create Dollar Journal // //"
            
            
            
            
def create_Dollar_Journals_Taxes(pivot,end_date,reversal,quarter_to_calculate,taxes):
    try:
        entities=get_entities()
        jrnl={} 
        Year_To_calculate = int(end_date.strftime("%Y"))
        year = f"{Year_To_calculate}"[:-2]
        Base_Bach_Name= f"Q{quarter_to_calculate}'{Year_To_calculate} - Bonus Accrual"
        Batch_Name = Base_Bach_Name
        Batch_Description = Base_Bach_Name
        #print(pivot) 
        currency="USD"
        period = end_date.strftime("%b-%y")
        count=0
        for i in pivot.keys():
            #print(i)
            data = entities[f"{i}"]
            group_id= datetime.today().strftime('%Y%m%d%H%M%S')
            Ledger_Country = data["Name"].split(" ")[1]
            Journal_Name = f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_Description =  f"{Base_Bach_Name} - {Ledger_Country} {group_id}"
            Line_Description =  f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_name = data["Name"]
            Ledger_id=data["Ledger"]
            Accounting_date = end_date.strftime("%Y-%m-%d")
            #datetime.datetime.now().strftime("%m/%d/%Y")
            Source = "AA Bot"
            Category  = "COMPandBEN_Accrual"
            entity = i
            #Line_Description = entities[f"first_data[2]"]
            entity_total=0
            header_soap = Create_SOAP_HEADER(Batch_Name,Ledger_id,period,Accounting_date)
            journal_line=""
            for j in pivot[i].keys():
                dpt=j
                account="60810"
                prod="000000"
                proj="000000"
                ico="00"
                fut="0"
                from_currency= pivot[i][j]["Currency"]
                convertion_rate = get_currency_exchange(from_currency,currency,end_date)
                debit= round(float(pivot[i][j]["Amount"]) * float(convertion_rate) *  float(taxes[i][quarter_to_calculate-1]),'.2f')
                print(debit)
                entity_total = entity_total+debit
                #print(currency,amount,entity_total)
                journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, debit, reversal, account, entity, dpt, Journal_Name,Ledger_name,Ledger_Description,0)
                jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Acounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : debit,"Credit":"0", "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description }
                count += 1
            dpt = "000"
            account="21110"
            credit=entity_total
            print(credit)
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, credit, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Acounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":credit, "Reversal" : reversal,"Acount" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description}
            count += 1
            foot= Create_SOAP_Foot()
            journal_entry = f"{header_soap}{journal_line}{foot}"
            response_to_push = Push_To_Oracle(journal_entry)
            if ">0</result>" not in response_to_push.text:
                print(response_to_push.text)
                print("Error on soap")
                return "Error pushing data converted taxes// //"
            importing = Create_Import_SOAP(set_id, source_id,Ledger_id , f"{group_id}")
            response_import = Import_to_GL(importing)
            import_id = json.loads(response_import.text)["ReqstId"]
            print(import_id)
            if import_id != -1:
                while True:
                    Status = json.loads(Get_Status(import_id).text)["items"][0]["RequestStatus"]
                    print(Status)
                    if Status == "SUCCEEDED":
                        print("Imported")
                        break
                    if Status == "ERROR" or Status == "WARNING":
                        print("Error on import")
                        return f"Python Error: Warning or Error on Importing PID : {import_id}   Ledger ID: {Ledger_id} on concerted taxes // //"
                        #return "Python Error: Warning or Error on Importing"
            else:
                print("Oracle's Error on import")
                return "Error on import journal converted taxes// //"
            with open(f"{data_path}\\Dollar_journal_taxes.txt","a+") as jrn:
                jrn.write(journal_entry)
        return jrnl
    except:
        return "Error cant create Taxes Dollar Journal  // //"





'''

        
        
    
def Create_SOAP_HEADER(Batch_Name, ledger_id, period_name, accounting_date):
    return f'''<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:typ="http://xmlns.oracle.com/apps/financials/generalLedger/journals/desktopEntry/journalImportService/types/" xmlns:jour="http://xmlns.oracle.com/apps/financials/generalLedger/journals/desktopEntry/journalImportService/" xmlns:jour1="http://xmlns.oracle.com/apps/flex/financials/generalLedger/journals/desktopEntry/journalLineGdf/">
   <soapenv:Header/>
   <soapenv:Body>
      <typ:importJournals>
         <typ:interfaceRows>
            <jour:BatchName>{Batch_Name}</jour:BatchName>
            <jour:BatchDescription>{Batch_Name}</jour:BatchDescription>
            <jour:LedgerId>{ledger_id}</jour:LedgerId>
            <jour:AccountingPeriodName>{period_name}</jour:AccountingPeriodName>
            <jour:AccountingDate>{accounting_date}</jour:AccountingDate>
            <jour:UserSourceName>AA Bot</jour:UserSourceName>
            <jour:UserCategoryName>COMPandBEN_Accrual</jour:UserCategoryName>
            <jour:ErrorToSuspenseFlag>True</jour:ErrorToSuspenseFlag>
            <jour:SummaryFlag>True</jour:SummaryFlag>
            <jour:ImportDescriptiveFlexField>N</jour:ImportDescriptiveFlexField>
'''


def Create_SOAP_Body(ledger_id, period_name, accounting_date, group_id, currency, amount, reversal, account, ent, dep, Journal_name,ledger_name,Ledger_Description,type):
    if type==0:
        crdr = "Dr"
    if type == 1:
        crdr = "Cr"
    return f'''               <jour:GlInterface>
               <jour:LedgerId>{ledger_id}</jour:LedgerId>
               <jour:LedgerName>{ledger_name}</jour:LedgerName>
               <jour:PeriodName>{period_name}</jour:PeriodName>
               <jour:AccountingDate>{accounting_date}</jour:AccountingDate>
               <jour:UserJeSourceName>AA BOT</jour:UserJeSourceName>
               <jour:UserJeCategoryName>COMPandBEN_Accrual</jour:UserJeCategoryName>
               <jour:GroupId>{group_id}</jour:GroupId>
               <jour:ChartOfAccountsId/>
               <jour:BalanceType>A</jour:BalanceType>
               <jour:CodeCombinationId/>
               <jour:Segment1>{ent}</jour:Segment1>
               <jour:Segment2>{dep}</jour:Segment2>
               <jour:Segment3>{account}</jour:Segment3>
               <jour:Segment4>000000</jour:Segment4>
               <jour:Segment5>000000</jour:Segment5>
               <jour:Segment6>00</jour:Segment6>
               <jour:Segment7>0</jour:Segment7>
               <jour:CurrencyCode>{currency}</jour:CurrencyCode>
               <jour:Entered{crdr}Amount currencyCode="{currency}">{amount}</jour:Entered{crdr}Amount>
               <jour:AccountedCr/>
               <jour:AccountedDr/>
               <jour:UserCurrencyConversionType>User</jour:UserCurrencyConversionType>
               <jour:CurrencyConversionDate>{accounting_date}</jour:CurrencyConversionDate>
               <jour:CurrencyConversionRate>1</jour:CurrencyConversionRate>	
               <jour:Reference4>{Ledger_Description}</jour:Reference4> 
               <jour:Reference5>{Ledger_Description}</jour:Reference5>
               <jour:Reference7>Y</jour:Reference7>          
               <jour:Reference8>{reversal}</jour:Reference8>       
               <jour:Reference9>Y</jour:Reference9>
            </jour:GlInterface>\n'''
            
def Create_SOAP_Foot():
    return '''         </typ:interfaceRows>
      </typ:importJournals>
   </soapenv:Body>
</soapenv:Envelope>\n
'''





def Push_To_Oracle(soap):
    url = f"{oracle_url}/fscmService/JournalImportService?WSDL"
    headerss = {"Content-Type": "text/xml;charset=UTF-8",
                "Accept-Encoding": "gzip, deflate",
                "Connection": "Keep-Alive"
                }
    logger.info("Read credentials")
    user,passw  = read_credentials()
    return requests.post(url, data=soap, auth=(user, passw), headers=headerss)



def AutoPost(soap):
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw  = read_credentials()
    #print(user,passw)
    return requests.post(url, data=soap, auth=(user, passw), headers=headerss)


def Get_Status(RqstId):
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations?finder=ESSJobStatusRF;requestId={RqstId}"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw  = read_credentials()
    return requests.get(url, auth=(user, passw), headers=headerss)


def Create_AutoPost():
    return f'''{{
    "OperationName":"submitESSJobRequest",
    "JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common/",
    "JobDefName":"AutomaticPosting",
    "ESSParameters":"{Autopost_id}",
    "ReqstId":null
    }} '''
    
def Create_Import_SOAP(set_id, source_id, ledger_id, group_id):
    return f'''{{ "OperationName":"submitESSJobRequest",
    "JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common/",
    "JobDefName":"JournalImportLauncher",
    "ESSParameters":"{set_id},{source_id},{ledger_id},{group_id},N,N,N",
    "ReqstId":null
    }}
    '''
    
def Import_to_GL(soap):
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw  = read_credentials()
    #print(user,passw)
    return requests.post(url, data=soap, auth=(user, passw), headers=headerss)




def Get_Ledger_Status(Period):
    #print(Period)
    entities = get_entities()
    flg=0
    String = f"Error : Next Ledgers are closed for period {Period} : <br>"
    try:
        for i in entities.keys():
            ledger_status = verify_closed_periods(Period,entities[i]["Ledger"])
            ledger_json=json.loads(ledger_status.text)
            ledger_json = ledger_json["items"][0]
            #print(ledger_json)
            if ledger_json["ClosingStatus"] == "C":
                flg=1
                closed = entities[i]["Name"]
                String += f"{closed}<br>"
        if flg ==0:
            return "0"
        else:
            return f"Error : {String}"
    except:
        return "Error : Not Well formed data // // "
    
    
    
def verify_closed_periods(Period,Ledger_id):
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/accountingPeriodStatusLOV?q=PeriodNameId={Period};ApplicationId=101;LedgerId={Ledger_id}"
    #print(url)
    user,passw  = read_credentials()
    headerss = {"Content-Type": "application/json",
            "Connection": "Keep-Alive"
            }    
    return requests.get(url, auth=(user, passw), headers=headerss)



def verify(data):
    logger.info(f"Trying to verify data {data}")
    data=data.upper()
    try:
        if "OVERRIDE" in data:
            return "ok"
        else:
            try:
                logger.info("Reading Parameters")
                XLS = read_xls_parameters()
                if type(XLS) == str:
                        return f"Error : Parameter file cant be read"
            except:
                    return "Error : Parameter file cant be read"
            else:      
                    start_date = XLS.iloc[0,1]
                    quarter_to_calculate = int(XLS.iloc[0,5])
                    Year_To_calculate = int(start_date.strftime("%Y")) + 1
                    data =  f"{Year_To_calculate}-{quarter_to_calculate}"
                    try:
                        with open(f"{config_path}\\config.cfg", "r") as config:
                            lst = config.read().split("\n")
                            if data in lst:
                                return ("Python Error: This Year-Quarter has been already running and override has not been defined ")
                    except:
                        with open(f"{config_path}\\config.cfg", "a+") as config:
                            config.write(data)
                            return "Ok"
                    
    except:
        logger.info(f"Data received not well defined")
        logger.info(f"Trying to push the log into the s3_bucket")
        Push_To_S3(f"{logs_path}\\functions.log", "Process4", "Log")
        return "Python Error: Region, Period or Ledger not well defined."



def Push_To_S3(File, process, subdir):
    '''
    This function push a file into the S3 bucket
    '''
    try:
        with open(File, "rb") as f:
            key = File.split("\\")[-1]
            response = s3_bucket.upload_fileobj(f, bucket_name, f"{process}/{subdir}/{key}")
            f.close()
    except Exception as e:
        logger.info(f"Error occurs while uploading")
        return "Python Error: Error occurs while uploading <BR>"
    
    

    
def clear():
    files_to_delete = os.listdir(data_path)
    for item in files_to_delete:
        if item.endswith(".xlsx"):
            os.remove(os.path.join(data_path, item))
    for item in files_to_delete:
        if item.endswith(".csv"):
            os.remove(os.path.join(data_path, item))
    for item in files_to_delete:
        if item.endswith(".xml"):
            os.remove(os.path.join(ROOT_DIR, item))