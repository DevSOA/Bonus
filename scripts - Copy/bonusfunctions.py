import csv
import os
import pandas as pd
import json
import logzero
from datetime import datetime
from calendar import isleap
import sys
import os
import xml.etree.ElementTree as ET
import csv
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import xmltodict
from io import BytesIO
from operator import itemgetter
import json
from io import BytesIO
from pdfdocument.document import PDFDocument
import json
import csv
import requests
from dateutil.relativedelta import relativedelta
from logzero import logger
import base64
from decimal import Decimal
import boto3
from datetime import datetime, timedelta
from lxml import etree
from collections import deque
import time

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
logs_path = os.path.join(ROOT_DIR, "Logs")
os.makedirs(logs_path, exist_ok=True)
data_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(data_path, exist_ok=True)
config_path = os.path.join(ROOT_DIR, "Config")
os.makedirs(config_path, exist_ok=True)
logzero.logfile(f"{logs_path}\\functions.log")
logzero.loglevel(logzero.INFO)
general_data_path = os.path.join(ROOT_DIR, "..\\generaldata")
os.makedirs(general_data_path, exist_ok=True)
logzero.json()
#oracle_url="https://efow.fs.us2.oraclecloud.com"
oracle_url="https://efow-test.fa.us2.oraclecloud.com"
#oracle_url= "https://efow-dev1.fa.us2.oraclecloud.com"

set_ndigits=6


set_id = "300000001414463"
Autopost_id = "300000150481649"
source_id = "300000150481632"
######## DEV DATA
#Autopost_id = "300000146748162"
#source_id= "300000146748158"

s3_bucket = boto3.client('s3')
lambda_call = boto3.client('lambda', region_name='us-west-2')

bucket_name= "gwre-rpa-test"


QUARTER1S="01/08/2020"
QUARTER2S="01/08/2020"
QUARTER3S="01/08/2020"
QUARTER4S="01/08/2020"

months = (["Jan", "Feb", "Mar", "Apr",
           "May", "Jun", "Jul", "Aug",
           "Sep", "Oct", "Nov", "Dec"]
          )


   
def set_user(data):
    logger.info("Setting credentials")
    try:
       user,passw= data.split("///")
       with open(f"{ROOT_DIR}\\Data\\credentials.txt","wb") as cred:
            cred.write(base64.b64encode(user.encode()))
            cred.write(b"\n")
            cred.write(base64.b64encode(passw.encode()))
            cred.close()
    except:
        logger.info("Error on Setting credentials")
        return "Error on set credentials // //"


def read_credentials():
    try:
        logger.info("Reading credentials")
        usr_psw = open(f"{ROOT_DIR}\\Data\\credentials.txt","rb")
        user = base64.b64decode(usr_psw.readline().decode()).decode()
        passw = base64.b64decode(usr_psw.readline().decode()).decode()
        usr_psw.close()
        user = user.strip()
        passw = passw.strip()
        #user,passw="vsethi","Welcome1"
    except:
        logger.info("Error on read credentials")
        return "Error on read credentials // //"
    return user,passw
    
    
def set_WD(data):
    logger.info("Setting WD credentials")
    try:
       user,passw= data.split("///")
       with open(f"{ROOT_DIR}\\Data\\WD.txt","wb") as cred:
            cred.write(base64.b64encode(user.encode()))
            cred.write(b"\n")
            cred.write(base64.b64encode(passw.encode()))
            cred.close()
    except:
        logger.info("Error on Setting WD credentials")
        return "Error on set credentials // //"



def read_WD_credentials():
    try:
        logger.info("Reading WD credentials")
        usr_psw = open(f"{ROOT_DIR}\\Data\\WD.txt","rb")
        user = base64.b64decode(usr_psw.readline().decode()).decode()
        passw = base64.b64decode(usr_psw.readline().decode()).decode()
        usr_psw.close()
    except:
        logger.info("Error on reading credentials")
        return "Error on read credentials // //"
    return user,passw



def get_ledgers_info():
    try:
        logger.info("Getting ledgers info to look for closed ones")
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/ledgersLOV?limit=10000"
        #print(url)
        headerss = {"Content-Type": "application/json",
                "Connection":  "Keep-Alive"
                }
        user,passw  = read_credentials()
    except:
        logger.info("Error on reading ledgers info|")
        return "Error on read ledger info // //"
    return requests.get(url, auth=(user, passw), headers=headerss)



''''''
def get_entities():
    logger.info("Reading Ledgers Id from Oracle")
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        #print(ledgers_info)
        #print(ledgers_info.text)
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        for i in ledger_json:
            ledgers[i["Name"]] = i["LedgerId"]
    except:
        logger.info("Error on getting entities")
        return "Error on set the entities // // "
    return  {
        "10": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "12": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "14": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "15": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "16": {"Name": "GW CA PL", "Ledger":  ledgers["GW CA PL"]},
        "65": {"Name": "GW BR PL", "Ledger":  ledgers["GW BR PL"]},
        "41": {"Name": "GW UK PL", "Ledger":  ledgers["GW UK PL"]},
        "43": {"Name": "GW FR PL", "Ledger":  ledgers["GW FR PL"]},
        "45": {"Name": "GW DE PL", "Ledger":  ledgers["GW DE PL"]},
        "46": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "47": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "48": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "49": {"Name": "GW IT PL", "Ledger":  ledgers["GW IT PL"]},
        "51": {"Name": "GW PL PL", "Ledger":  ledgers["GW PL PL"]},
        "55": {"Name": "GW CH PL", "Ledger":  ledgers["GW CH PL"]},
        "57": {"Name": "GW ES PL", "Ledger":  ledgers["GW ES PL"]},
        "71": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "72": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "80": {"Name": "GW CN PL", "Ledger":  ledgers["GW CN PL"]},
        "86": {"Name": "GW JP PL", "Ledger":  ledgers["GW JP PL"]},
        "84": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "85": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "82": {"Name": "GW MY PL", "Ledger":  ledgers["GW MY PL"]},
        "61": {"Name": "GW AR PL", "Ledger":  ledgers["GW AR PL"]},
        "53": {"Name": "GW AT PL", "Ledger":  ledgers["GW AT PL"]},
        "58": {"Name": "GW DK PL", "Ledger":  ledgers["GW DK PL"]}, 
    }
''''''


def get_entities(ledger_type):
    logger.info(f"Getting entities")
    logger.info("Reading Ledgers Id from Oracle")
    if ledger_type.upper().strip() == "PRIMARY":
        ledger_file=f"{general_data_path}\\primary.txt"
    elif ledger_type.upper().strip() == "SUBLEDGER":
        ledger_file=f"{general_data_path}\\subledgers.txt"
    with open(ledger_file,"r") as ledger_config:
        ledger_list=ledger_config.read().splitlines()
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        for i in ledger_json:
            ledgers[i["Name"]] = i["LedgerId"]        
        entities={}
        for i in ledger_list:
            #print(i)
            try:
                data = i.split(",")
                #print(data)
                entity=data[0].strip()
                ledger_name = data[1].strip()
                entities.update({ entity: {"Name": ledger_name, "Ledger":  ledgers[ledger_name]} })  
            except:
                pass
    except:
        logger.info("Error on set the entities")
        return "Error on set the entities // // "
    return  entities


'''
entities = {
                        "10": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "12": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "14": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "15": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "16": {"Name": "GW CA PL", "Ledger": "300000001414017"},
                        "65": {"Name": "GW BR PL", "Ledger": "300000001414018"},
                        "41": {"Name": "GW UK PL", "Ledger": "300000001414019"},
                        "43": {"Name": "GW FR PL", "Ledger": "300000001414020"},
                        "45": {"Name": "GW DE PL", "Ledger": "300000001414021"},
                        "46": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "47": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "48": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "49": {"Name": "GW IT PL", "Ledger": "300000001414023"},
                        "51": {"Name": "GW PL PL", "Ledger": "300000001414024"},
                        "55": {"Name": "GW CH PL", "Ledger": "300000001414025"},
                        "57": {"Name": "GW ES PL", "Ledger": "300000001414026"},
                        "71": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "72": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "80": {"Name": "GW CN PL", "Ledger": "300000001414029"},
                        "86": {"Name": "GW JP PL", "Ledger": "300000001414030"},
                        "84": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "85": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "82": {"Name": "GW MY PL", "Ledger": "300000019642549"},
                        "61": {"Name": "GW AR PL", "Ledger": "300000019833039"},
                        "53": {"Name": "GW AT PL", "Ledger": "300000017944094"},
                        "58": {"Name": "GW DK PL", "Ledger": "300000026789264"},
                        }
'''
'''

def get_entities():
    return {
                        "10": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "12": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "14": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "15": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "16": {"Name": "GW CA PL", "Ledger": "300000001414017"},
                        "65": {"Name": "GW BR PL", "Ledger": "300000001414018"},
                        "41": {"Name": "GW UK PL", "Ledger": "300000001414019"},
                        "43": {"Name": "GW FR PL", "Ledger": "300000001414020"},
                        "45": {"Name": "GW DE PL", "Ledger": "300000001414021"},
                        "46": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "47": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "48": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "49": {"Name": "GW IT PL", "Ledger": "300000001414023"},
                        "51": {"Name": "GW PL PL", "Ledger": "300000001414024"},
                        "55": {"Name": "GW CH PL", "Ledger": "300000001414025"},
                        "57": {"Name": "GW ES PL", "Ledger": "300000001414026"},
                        "71": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "72": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "80": {"Name": "GW CN PL", "Ledger": "300000001414029"},
                        "86": {"Name": "GW JP PL", "Ledger": "300000001414030"},
                        "84": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "85": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "82": {"Name": "GW MY PL", "Ledger": "300000019642549"},
                        "61": {"Name": "GW AR PL", "Ledger": "300000019833039"},
                        "53": {"Name": "GW AT PL", "Ledger": "300000017944094"},
                        "58": {"Name": "GW DK PL", "Ledger": "300000026789264"},
                        }
'''

def Create_Workday_Url(Workday_String_Start_date,String_Cutoff,format):
    logger.info("Creating Workday Endpoint")
    try:
        percent = "%"
        url = f"https://wd5-services1.myworkday.com/ccx/service/customreport2/guidewire/Guidewire_ISU/Bonus_Calculation_Raas?Start_Date={Workday_String_Start_date}-08{percent}3A00&End_Date={String_Cutoff}-08{percent}3A00&format={format}"
        #print(f"URL :   {url}")
    except:
        logger.info("Error on creating url")
        return "Error: Error on url <br> // // "
    else:
        return url
    
    
def Save_Sheet(Dic, File, Sheet, Col, Row):
    logger.info("Saving Data in Xlsx")
    try:
        df = pd.DataFrame.from_dict(Dic).T
    except:
        logger.info("Error on saving Xlsx")
        return "Error on get the dictionary // // "
    else:
        try:
            book = load_workbook(File)
        except:
            book = openpyxl.Workbook()
            book.save(File)
            book = load_workbook(File)
        logger.info("Writing Xlsx data")
        writer = pd.ExcelWriter(File, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=Sheet, startcol=Col, startrow=Row)
        writer.save()
        
            
        
def read_xls_parameters():
    logger.info("Reading Parameters")
    try:
        df = pd.read_excel (f'{ROOT_DIR}\\Data\\Bonus_Parameters.xlsx')
        return df
    except:
        logger.info("Error on reading the excel file")
        return "Error : Cannot read the excel file // // "

        
        
def get_quarters(Year):
    logger.info("Creating predefined quarters")
    try:
        Prior_Year = int(Year)-1
        QUARTER1S=f"08/01/{Prior_Year}"
        QUARTER2S=f"11/01/{Prior_Year}"
        QUARTER3S=f"02/01/{Year}"
        QUARTER4S=f"05/01/{Year}"
        QUARTER1E=f"10/31/{Prior_Year}"
        QUARTER2E=f"01/31/{Year}"
        QUARTER3E=f"04/30/{Year}"
        QUARTER4E=f"07/31/{Year}"
    except:
        logger.info("Error on set the quarters")
        return "Error on set the quarters // // "
    return [(datetime.strptime(QUARTER1S,"%m/%d/%Y"),datetime.strptime(QUARTER1E,"%m/%d/%Y")) ,(datetime.strptime(QUARTER2S,"%m/%d/%Y"),datetime.strptime(QUARTER2E,"%m/%d/%Y")), (datetime.strptime(QUARTER3S,"%m/%d/%Y"),datetime.strptime(QUARTER3E,"%m/%d/%Y")), (datetime.strptime(QUARTER4S,"%m/%d/%Y"),datetime.strptime(QUARTER4E,"%m/%d/%Y"))]

def create_taxes(XLS):
    logger.info("Reading Taxes Info")
    try:
        taxes_dic={}
        for i in range(4,len(XLS["Bonus_Parameters"])):
            taxes_dic[XLS.iloc[i,0]]=[round(Decimal(XLS.iloc[i,1]),set_ndigits),round(Decimal(XLS.iloc[i,2]),set_ndigits),round(Decimal(XLS.iloc[i,3]),set_ndigits),round(Decimal(XLS.iloc[i,4]),set_ndigits)]
    except:
        logger.info("Error on create taxes")
        return "Error on create taxes // // "
    return taxes_dic
            

def get_employee_data_file(url,year,format):
    logger.info("Reading WorkDay data and saving in file")
    try:
        with open(f"{data_path}\\{year}_bonusfile.{format}", "w+b") as xml_to_write:
            user,passw = read_WD_credentials()
            #user="SVC_RPA_WDAY"
            #passw  = "Gu1d3wire!"
            bonus_data = requests.get(f"{url}", auth=(
                                                user,
                                                passw
                                                ),
                                            stream=True).content
            xml_to_write.write(bonus_data)
            return "0"
    except:
        logger.info("Error on saving workday data")
        return "Error when trying to save the workday data // // "
    
    
 
 
 
 
    
def Get_Xml_Data(year):
    logger.info("Reading xml file")
    try:    
        Xml_Data = open(f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml","r").read()  
        dictionary = xmltodict.parse(Xml_Data)
        json_data = json.dumps(dictionary)
        data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
        employee_data={}
        count=0
        Sec = 0
        Status = "1"
        Type = "Regular"
        HRDate= "2012-12-12" 
        PFTime="Full"
        Salary=float(0.0)
        Currency = "USD"
        Bonus_Plan = "Fixed Percent"
        Percent = 0.0
        Commission = 0.0
        CompID = 10
        Dpt = 101 
        Region = "AMER"
        PayG="GuideWire"
        for i in (data):
            try:
                    #print(i)
                    try:
                        Sec=int(i["wd:Row_Sequence"])
                        logger.info(f"Extracting Employee data for {Sec}")
                    except:
                        #print(f"RS Is not comming")
                        pass
                    try:
                        Status = i["wd:Active_Status"]
                    except:
                        #print(f"Act Is not comming")
                        pass        
                    try:
                        Type = i["wd:Employee_Type"]
                    except:
                        #print(f"Type Is not comming")
                        pass
                    try:
                        HRDate = i["wd:Hire_Date"]
                    except:
                        #print(f"HR Is not comming")
                        pass
                    try:
                        PFTime = i["wd:Time_Type"]['@wd:Descriptor']
                    except:
                        #print(f"Descriptor type Is not comming")
                        pass
                    try:
                        Salary = float(i["wd:Salary"])
                    except:
                        #print(f"Salary Is not comming")
                        pass
                    try:
                        Currency = i["wd:Salary_Currency"]
                    except:
                        #print(f"Currency Is not comming")
                        pass
                    try:
                        Bonus_Plan = i["wd:Bonus_Plan_Type"]
                    except:
                        #print(i)
                        #print(f"BP Is not comming on employee {Sec}")
                        pass
                    try:
                        Percent = float(i["wd:Bonus_Percent"])
                    except:
                        #print(f"Perc. Is not comming")
                        pass
                    try:
                        Commission = float(i["wd:Commission_Amount"])    
                    except:
                        #print(f"Comm Is not comming")
                        pass
                    try:
                        CompID = i["wd:Company_ID"]    
                    except:
                        #print(f"ID COmp Is not comming")
                        pass
                    try:
                        Dpt = i["wd:Cost_Center"]
                    except:
                        #print(f"DPT Is not comming")
                        pass
                    try:
                        Region = i["wd:Region"]    
                    except:
                        #print(f"Region Is not comming")
                        pass
                    try:
                        PayG=i["wd:Pay_Group"]
                    except:
                        #print(f"PG Is not comming")
                        pass
                    employee_data[count]={
                    "Employee_ID":Sec,
                    "Employee_Status":Status,
                    "Employee_Type":Type,
                    "Employee_HE_Date":HRDate,
                    "Employee_PF_Time":PFTime,
                    "Employee_Current_Salary":Salary,
                    "Employee_Current_Currency":Currency,
                    "Employee_Current_Bonus_Plan":Bonus_Plan,
                    "Employee_Current_Bonus_Percent":Percent,
                    "Employee_Current_Commission":Commission,
                    "Employee_Current_Company_ID":CompID,
                    "Employee_Current_DPT":Dpt,
                    "Employee_Current_Region":Region,
                    "Employee_Current_Pay_Group":PayG}
                    count+=1
                    try:
                        if type((i)['wd:Compensation_History']) is dict:
                            try:
                                #HRDate = i["wd:Compensation_History"]["wd:Effective_Date_Compensation"]
                                HRDate =  employee_data[count-1]["Employee_HE_Date"]
                                employee_data[count-1]["Employee_HE_Date"] = i["wd:Compensation_History"]["wd:Effective_Date_Compensation"]
                            except:
                                #print(f"COMP HR Is not comming\n\n\n\n\n\n\n\n\n\n\n")
                                pass
                            try:
                                Salary = i["wd:Compensation_History"]["wd:Prior_Salary"]
                            except:
                                #print(f"COMP SAL Is not comming")
                                pass
                            try:
                                Currency = i["wd:Compensation_History"]["wd:Prior_Salary_Currency"]
                            except:
                                #print(f"COMP CURR Is not comming")
                                pass
                            try:
                                Percent = i["wd:Compensation_History"]["wd:Prior_Bonus_Percent"]
                            except:
                                #print(f"COMP BONUS PERC Is not comming")
                                pass
                            employee_data[count]={
                            "Employee_ID":Sec,
                            "Employee_Status":Status,
                            "Employee_Type":Type,
                            "Employee_HE_Date":HRDate,
                            "Employee_PF_Time":PFTime,
                            "Employee_Current_Salary":Salary,
                            "Employee_Current_Currency":Currency,
                            "Employee_Current_Bonus_Plan":Bonus_Plan,
                            "Employee_Current_Bonus_Percent":Percent,
                            "Employee_Current_Commission":Commission,
                            "Employee_Current_Company_ID":CompID,
                            "Employee_Current_DPT":Dpt,
                            "Employee_Current_Region":Region,
                            "Employee_Current_Pay_Group":PayG}
                            #print("Dcitionary --->",HRDate,employee_data[count-1]["Employee_HE_Date"],employee_data[count]["Employee_HE_Date"])
                            #employee_data[count]['Employee_HE_Date'] = nHRDate
                            count+=1
                        elif type((i)['wd:Compensation_History']) is list:
                            datelist=[HRDate]
                            for j in range(len(i['wd:Compensation_History'])):
                                try:
                                    HRDate =   i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"]
                                    datelist.append(HRDate)
                                    #HRDate = i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"]
                                except:
                                    #print(f"SEC COMP HR Is not comming\n\n\n\n\n\n\n\n\n\n\n")
                                    pass
                                try:
                                    Salary = i["wd:Compensation_History"][j]["wd:Prior_Salary"]
                                except:
                                    #print(f"SEC COMP SAL Is not comming")
                                    pass
                                try:
                                    Currency = i["wd:Compensation_History"][j]["wd:Prior_Salary_Currency"]
                                except:
                                    #print(f"SEC COMP CURR Is not comming")
                                    pass
                                try:
                                    Percent = i["wd:Compensation_History"][j]["wd:Prior_Bonus_Percent"]
                                except:
                                    #print(f"SEC COMP PERC Is not comming")
                                    pass
                                employee_data[count]={
                                "Employee_ID":Sec,
                                "Employee_Status":Status,
                                "Employee_Type":Type,
                                "Employee_HE_Date":HRDate,
                                "Employee_PF_Time":PFTime,
                                "Employee_Current_Salary":Salary,
                                "Employee_Current_Currency":Currency,
                                "Employee_Current_Bonus_Plan":Bonus_Plan,
                                "Employee_Current_Bonus_Percent":Percent,
                                "Employee_Current_Commission":Commission,
                                "Employee_Current_Company_ID":CompID,
                                "Employee_Current_DPT":Dpt,
                                "Employee_Current_Region":Region,
                                "Employee_Current_Pay_Group":PayG}
                                count+=1
                            #employee_data[count-1-len(datelist)]['Employee_HE_Date'] = datelist[-1]
                            #for j in range(1,len(datelist)):
                            #    employee_data[count-1-len(datelist)+j]['Employee_HE_Date'] = datelist[j-1]
                            #temp = employee_data[count-counterdates-1]["Employee_Current_Salary"]
                            #for j in range(counterdates,0,-1):
                            #    #employee_data[count-1-j]["Employee_HE_Date"] =
                            #    back = j+1
                            #    print(back,j,count,count - back)
                            #    employee_data[count-back]["Employee_Current_Salary"] = employee_data[count-j]["Employee_Current_Salary"]
                            #employee_data[count-1]["Employee_Current_Salary"] =  temp                    
                            #print("list--->",HRDate,employee_data[count-2]["Employee_HE_Date"],employee_data[count-1]["Employee_HE_Date"])
                            #for j in range(len(datelist)):
                            #    employee_data[count-1-j]['Employee_HE_Date']=datelist[j]
                            #employee_data[count-1-len(datelist)]['Employee_HE_Date'] = datelist[0]
                            #employee_data[count-len(datelist)]['Employee_HE_Date'] = datelist[-1]
                            #employee_data[count-1]['Employee_HE_Date'] = datelist[-1]
                            #for j in range(len(datelist),0,-1):
                            #    employee_data[count-1-j]['Employee_HE_Date'] = datelist[len(datelist)-1-j]
                            #employee_data[count-1-len(datelist)]['Employee_HE_Date'] = datelist[-1]
                            #print(datelist)
                            #####for k in range(0,len(datelist)):
                            #####    print(employee_data[count-len(datelist)+k]['Employee_HE_Date'],employee_data[count-len(datelist)+k]['Employee_Current_Salary'])
                            employee_data[count-len(datelist)]['Employee_HE_Date'] = datelist[-1]
                            for k in range(1,len(datelist)):
                                employee_data[count-len(datelist)+k]['Employee_HE_Date'] = datelist[k-1]
                            #print("\n")
                            #for k in range(0,len(datelist)):
                            #    print(employee_data[count-len(datelist)+k]['Employee_HE_Date'],employee_data[count-len(datelist)+k]['Employee_Current_Salary'],datelist[k-1])
                    except :
                        pass
            except:
                    #print("Error on getting values\n")
                    pass
    except: 
        logger.info("Error reading xml")
        return "Error at time to read the XML file // // "
    else:
        logger.info("Saving employee data")
        #Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Data_From_WorkDay", 0, 0)
        return employee_data
 
    
    
'''    
def Get_Xml_Data(year):
    logger.info("Reading xml file")
    try:    
        Xml_Data = open(f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml","r").read()  
        dictionary = xmltodict.parse(Xml_Data)
        json_data = json.dumps(dictionary)
        data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
        employee_data={}
        count=0
        Sec = 0
        Status = "1"
        Type = "Regular"
        HRDate= "2012-12-12" 
        PFTime="Full"
        Salary=float(0.0)
        Currency = "USD"
        Bonus_Plan = "Fixed Percent"
        Percent = 0.0
        Commission = 0.0
        CompID = 10
        Dpt = 101 
        Region = "AMER"
        PayG="GuideWire"
        try:
            logger.info("Extracting Employee data")
            for i in (data):
                    #print(i)
                    try:
                        Sec=int(i["wd:Row_Sequence"])
                    except:
                        print(f"RS Is not comming")
                    try:
                        Status = i["wd:Active_Status"]
                    except:
                        print(f"Act Is not comming")        
                    try:
                        Type = i["wd:Employee_Type"]
                    except:
                        print(f"Type Is not comming")
                    try:
                        HRDate = i["wd:Hire_Date"]
                    except:
                        print(f"HR Is not comming")
                    try:
                        PFTime = i["wd:Time_Type"]['@wd:Descriptor']
                    except:
                        print(f"Descriptor type Is not comming")
                    try:
                        Salary = float(i["wd:Salary"])
                    except:
                        print(f"Salary Is not comming")
                    try:
                        Currency = i["wd:Salary_Currency"]
                    except:
                        print(f"Currency Is not comming")
                    try:
                        Bonus_Plan = i["wd:Bonus_Plan_Type"]
                    except:
                        #print(i)
                        print(f"BP Is not comming on employee {Sec}")
                    try:
                        Percent = float(i["wd:Bonus_Percent"])
                    except:
                        print(f"Perc. Is not comming")
                    try:
                        Commission = float(i["wd:Commission_Amount"])    
                    except:
                        print(f"Comm Is not comming")
                    try:
                        CompID = i["wd:Company_ID"]    
                    except:
                        print(f"ID COmp Is not comming")
                    try:
                        Dpt = i["wd:Cost_Center"]
                    except:
                        print(f"DPT Is not comming")
                    try:
                        Region = i["wd:Region"]    
                    except:
                        print(f"Region Is not comming")
                    try:
                        PayG=i["wd:Pay_Group"]
                    except:
                        print(f"PG Is not comming")
                    employee_data[count]={
                    "Employee_ID":Sec,
                    "Employee_Status":Status,
                    "Employee_Type":Type,
                    "Employee_HE_Date":HRDate,
                    "Employee_PF_Time":PFTime,
                    "Employee_Current_Salary":Salary,
                    "Employee_Current_Currency":Currency,
                    "Employee_Current_Bonus_Plan":Bonus_Plan,
                    "Employee_Current_Bonus_Percent":Percent,
                    "Employee_Current_Commission":Commission,
                    "Employee_Current_Company_ID":CompID,
                    "Employee_Current_DPT":Dpt,
                    "Employee_Current_Region":Region,
                    "Employee_Current_Pay_Group":PayG}
                    count+=1
                    try:
                        if type((i)['wd:Compensation_History']) is dict:
                            try:
                                nHRDate = i["wd:Compensation_History"]["wd:Effective_Date_Compensation"]
                            except:
                                print(f"COMP HR Is not comming")
                            try:
                                Salary = i["wd:Compensation_History"]["wd:Prior_Salary"]
                            except:
                                print(f"COMP SAL Is not comming")
                            try:
                                Currency = i["wd:Compensation_History"]["wd:Prior_Salary_Currency"]
                            except:
                                print(f"COMP CURR Is not comming")
                            try:
                                Percent = i["wd:Compensation_History"]["wd:Prior_Bonus_Percent"]
                            except:
                                print(f"COMP BONUS PERC Is not comming")
                            employee_data[count]={
                            "Employee_ID":Sec,
                            "Employee_Status":Status,
                            "Employee_Type":Type,
                            "Employee_HE_Date":HRDate,
                            "Employee_PF_Time":PFTime,
                            "Employee_Current_Salary":Salary,
                            "Employee_Current_Currency":Currency,
                            "Employee_Current_Bonus_Plan":Bonus_Plan,
                            "Employee_Current_Bonus_Percent":Percent,
                            "Employee_Current_Commission":Commission,
                            "Employee_Current_Company_ID":CompID,
                            "Employee_Current_DPT":Dpt,
                            "Employee_Current_Region":Region,
                            "Employee_Current_Pay_Group":PayG}
                            employee_data[count-1]['Employee_HE_Date'] = nHRDate
                            count+=1
                        elif type((i)['wd:Compensation_History']) is list:
                                counterdates=0
                                for j in range(len(i['wd:Compensation_History'])):
                                    try:
                                        HRDate = i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"]
                                    except:
                                        print(f"SEC COMP HR Is not comming")
                                    try:
                                        Salary = i["wd:Compensation_History"][j]["wd:Prior_Salary"]
                                    except:
                                        print(f"SEC COMP SAL Is not comming")
                                    try:
                                        Currency = i["wd:Compensation_History"][j]["wd:Prior_Salary_Currency"]
                                    except:
                                        print(f"SEC COMP CURR Is not comming")
                                    try:
                                        Percent = i["wd:Compensation_History"][j]["wd:Prior_Bonus_Percent"]
                                    except:
                                        print(f"SEC COMP PERC Is not comming")
                                    employee_data[count]={
                                    "Employee_ID":Sec,
                                    "Employee_Status":Status,
                                    "Employee_Type":Type,
                                    "Employee_HE_Date":HRDate,
                                    "Employee_PF_Time":PFTime,
                                    "Employee_Current_Salary":Salary,
                                    "Employee_Current_Currency":Currency,
                                    "Employee_Current_Bonus_Plan":Bonus_Plan,
                                    "Employee_Current_Bonus_Percent":Percent,
                                    "Employee_Current_Commission":Commission,
                                    "Employee_Current_Company_ID":CompID,
                                    "Employee_Current_DPT":Dpt,
                                    "Employee_Current_Region":Region,
                                    "Employee_Current_Pay_Group":PayG}
                                    counterdates+=1
                                    count+=1
                                temp = employee_data[count-counterdates-1]["Employee_Current_Salary"]
                                for j in range(counterdates,0,-1):
                                    #employee_data[count-1-j]["Employee_HE_Date"] =
                                    back = j+1
                                    print(back,j,count,count - back)
                                    employee_data[count-back]["Employee_Current_Salary"] = employee_data[count-j]["Employee_Current_Salary"]
                                employee_data[count-1]["Employee_Current_Salary"] =  temp                    
                                
                    except :
                        pass
        except:
             print("Error on getting values\n")
             pass
    except: 
        logger.info("Error reading xml")
        return "Error at time to read the XML file // // "
    else:
        logger.info("Saving employee data")
        #Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Data_From_WorkDay", 0, 0)
        return employee_data
'''




 
    

    
'''
def Get_Xml_Data(year):
        logger.info("Reading xml file")
    #try:    
        Xml_Data = open(f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml","r").read()  
        dictionary = xmltodict.parse(Xml_Data)
        json_data = json.dumps(dictionary)
        data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
        employee_data={}
        count=0
        Sec = 0
        Status = "1"
        Type = "Regular"
        HRDate= "2012-12-12" 
        PFTime="Full"
        Salary=float(0.0)
        Currency = "USD"
        Bonus_Plan = "Fixed Percent"
        Percent = 0.0
        Commission = 0.0
        CompID = 10
        Dpt = 101 
        Region = "AMER"
        PayG="GuideWire"
        try:
            logger.info("Extracting Employee data")
            for i in (data):
                    #print(i)
                    try:
                        Sec=int(i["wd:Row_Sequence"])
                    except:
                        print(f"RS Is not comming")
                    try:
                        Status = i["wd:Active_Status"]
                    except:
                        print(f"Act Is not comming")        
                    try:
                        Type = i["wd:Employee_Type"]
                    except:
                        print(f"Type Is not comming")
                    try:
                        HRDate = i["wd:Hire_Date"]
                    except:
                        print(f"HR Is not comming")
                    try:
                        PFTime = i["wd:Time_Type"]['@wd:Descriptor']
                    except:
                        print(f"Descriptior type Is not comming")
                    try:
                        Salary = float(i["wd:Salary"])
                    except:
                        print(f"Salary Is not comming")
                    try:
                        Currency = i["wd:Salary_Currency"]
                    except:
                        print(f"Currency Is not comming")
                    try:
                        Bonus_Plan = i["wd:Bonus_Plan_Type"]
                    except:
                        #print(i)
                        print(f"BP Is not comming on employee {Sec}")
                    try:
                        Percent = float(i["wd:Bonus_Percent"])
                    except:
                        print(f"Perc. Is not comming")
                    try:
                        Commission = float(i["wd:Commission_Amount"])    
                    except:
                        print(f"Comm Is not comming")
                    try:
                        CompID = i["wd:Company_ID"]    
                    except:
                        print(f"ID COmp Is not comming")
                    try:
                        Dpt = i["wd:Cost_Center"]
                    except:
                        print(f"DPT Is not comming")
                    try:
                        Region = i["wd:Region"]    
                    except:
                        print(f"Region Is not comming")
                    try:
                        PayG=i["wd:Pay_Group"]
                    except:
                        print(f"PG Is not comming")
                    employee_data[count]={
                    "Employee_ID":Sec,
                    "Employee_Status":Status,
                    "Employee_Type":Type,
                    "Employee_HE_Date":HRDate,
                    "Employee_PF_Time":PFTime,
                    "Employee_Current_Salary":Salary,
                    "Employee_Current_Currency":Currency,
                    "Employee_Current_Bonus_Plan":Bonus_Plan,
                    "Employee_Current_Bonus_Percent":Percent,
                    "Employee_Current_Commission":Commission,
                    "Employee_Current_Company_ID":CompID,
                    "Employee_Current_DPT":Dpt,
                    "Employee_Current_Region":Region,
                    "Employee_Current_Pay_Group":PayG}
                    count+=1
                    try:
                        if type((i)['wd:Compensation_History']) is dict:
                            try:
                                HRDate = i["wd:Compensation_History"]["wd:Effective_Date_Compensation"]
                            except:
                                print(f"COMP HR Is not comming")
                            try:
                               Salary = i["wd:Compensation_History"]["wd:Proposed_Salary"]
                            except:
                                print(f"COMP SAL Is not comming")
                            try:
                               Currency = i["wd:Compensation_History"]["wd:Proposed_Salary_Currency"]
                            except:
                                print(f"COMP CURR Is not comming")
                            try:
                               Percent = i["wd:Compensation_History"]["wd:Proposed_Bonus_Percent"]
                            except:
                                print(f"COMP BONUS PERC Is not comming")
                            
                            employee_data[count]={
                            "Employee_ID":Sec,
                            "Employee_Status":Status,
                            "Employee_Type":Type,
                            "Employee_HE_Date":HRDate,
                            "Employee_PF_Time":PFTime,
                            "Employee_Current_Salary":Salary,
                            "Employee_Current_Currency":Currency,
                            "Employee_Current_Bonus_Plan":Bonus_Plan,
                            "Employee_Current_Bonus_Percent":Percent,
                            "Employee_Current_Commission":Commission,
                            "Employee_Current_Company_ID":CompID,
                            "Employee_Current_DPT":Dpt,
                            "Employee_Current_Region":Region,
                            "Employee_Current_Pay_Group":PayG}
                            count+=1
                        elif type((i)['wd:Compensation_History']) is list:
                                for j in range(len(i['wd:Compensation_History'])):
                                    try:
                                        HRDate = i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"]
                                    except:
                                        print(f"SEC COMP HR Is not comming")
                                    try:
                                        Salary = i["wd:Compensation_History"][j]["wd:Proposed_Salary"]
                                    except:
                                        print(f"SEC COMP SAL Is not comming")
                                    try:
                                        Currency = i["wd:Compensation_History"][j]["wd:Proposed_Salary_Currency"]
                                    except:
                                        print(f"SEC COMP CURR Is not comming")
                                    try:
                                        Percent = i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Percent"]
                                    except:
                                        print(f"SEC COMP PERC Is not comming")
                                    employee_data[count]={
                                    "Employee_ID":Sec,
                                    "Employee_Status":Status,
                                    "Employee_Type":Type,
                                    "Employee_HE_Date":HRDate,
                                    "Employee_PF_Time":PFTime,
                                    "Employee_Current_Salary":Salary,
                                    "Employee_Current_Currency":Currency,
                                    "Employee_Current_Bonus_Plan":Bonus_Plan,
                                    "Employee_Current_Bonus_Percent":Percent,
                                    "Employee_Current_Commission":Commission,
                                    "Employee_Current_Company_ID":CompID,
                                    "Employee_Current_DPT":Dpt,
                                    "Employee_Current_Region":Region,
                                    "Employee_Current_Pay_Group":PayG}
                                    count+=1
                    except :
                        pass
        except:
             print("Error on getting values\n")
             pass
    #except: 
    #     return "Error at time to read the XML file // // "
    #else:
        #Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Data_From_WorkDay", 0, 0)
        return employee_data


'''
'''

def Get_Xml_Data(year):
    logger.info("Reading xml file")
    try:    
        Xml_Data = open(f"{data_path}\\newxml.xml","r").read()  
        dictionary = xmltodict.parse(Xml_Data)
        json_data = json.dumps(dictionary)
        data = json.loads(json_data)["wd:Report_Data"]["wd:Report_Entry"]
        employee_data={}
        count=0
        try:
            logger.info("Extracting Employee data")
            for i in (data):
                    print(i)
                    employee_data[count]={
                    "Employee_ID":i["wd:Row_Sequence"],
                    "Employee_Status":i["wd:Active_Status"],
                    "Employee_Type":i["wd:Employee_Type"],
                    "Employee_HE_Date":i["wd:Hire_Date"],
                    "Employee_PF_Time":i["wd:Time_Type"]['type'][1],
                    "Employee_Current_Salary":i["wd:Salary"],
                    "Employee_Current_Currency":i["wd:Salary_Currency"],
                    "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                    "Employee_Current_Bonus_Amount":i["wd:Bonus_Amount"],
                    "Employee_Current_Bonus_Percent":i["wd:Bonus_Percent"],
                    "Employee_Current_Commission":i["wd:Commission_Amount"],
                    "Employee_Current_Company":i["wd:Company"],
                    "Employee_Current_Company_ID":i["wd:Company_ID"],
                    "Employee_Current_DPT":i["wd:Cost_Center"],
                    #"Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                    "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                    "Employee_Current_Region":i["wd:Region"],
                    "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                    count+=1
                    try:
                        if type((i)['wd:Compensation_History']) is dict:
                            employee_data[count]={
                            "Employee_ID":i["wd:Row_Sequence"],
                            "Employee_Status":i["wd:Active_Status"],
                            "Employee_Type":i["wd:Employee_Type"],
                            "Employee_HE_Date":i["wd:Compensation_History"]["wd:Effective_Date_Compensation"],
                            "Employee_PF_Time":i["wd:Time_Type"]['type'][1],
                            "Employee_Current_Salary":i["wd:Compensation_History"]["wd:Proposed_Salary"],
                            "Employee_Current_Currency":i["wd:Compensation_History"]["wd:Proposed_Salary_Currency"],
                            "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                            "Employee_Current_Bonus_Amount":i["wd:Compensation_History"]["wd:Proposed_Bonus_Amount"],
                            "Employee_Current_Bonus_Percent":i["wd:Compensation_History"]["wd:Proposed_Bonus_Percent"],
                            "Employee_Current_Commission":i["wd:Commission_Amount"],
                            "Employee_Current_Company":i["wd:Company"],
                            "Employee_Current_Company_ID":i["wd:Company_ID"],
                            "Employee_Current_DPT":i["wd:Cost_Center"],
                            #"Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                            "Employee_Current_Country_Alpha":i["wd:Country"]["wd:ID"][2]["#text"],
                            "Employee_Current_Region":i["wd:Region"],
                            "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                            count+=1
                        elif type((i)['wd:Compensation_History']) is list:
                                for j in range(len(i['wd:Compensation_History'])):
                                    employee_data[count]={
                                    "Employee_ID":i["wd:Row_Sequence"],
                                    "Employee_Status":i["wd:Active_Status"],
                                    "Employee_Type":i["wd:Employee_Type"],
                                    "Employee_HE_Date":i["wd:Compensation_History"][j]["wd:Effective_Date_Compensation"],
                                    "Employee_PF_Time":i["wd:Time_Type"]['type'][1],
                                    "Employee_Current_Salary":i["wd:Compensation_History"][j]["wd:Proposed_Salary"],
                                    "Employee_Current_Currency":i["wd:Compensation_History"][j]["wd:Proposed_Salary_Currency"],
                                    "Employee_Current_Bonus_Plan":i["wd:Bonus_Plan_Type"],
                                    "Employee_Current_Bonus_Amount":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Amount"],
                                    "Employee_Current_Bonus_Percent":i["wd:Compensation_History"][j]["wd:Proposed_Bonus_Percent"],
                                    "Employee_Current_Commission":i["wd:Commission_Amount"],
                                    "Employee_Current_Company":i["wd:Company"],
                                    "Employee_Current_Company_ID":i["wd:Company_ID"],
                                    "Employee_Current_DPT":i["wd:Cost_Center"],
                                    #"Employee_Current_Country_Name":i["wd:Country"]["@wd:Descriptor"],
                                    "Employee_Current_Country_Alpha":i["wd:Country"]["type"][2]["#text"],
                                    "Employee_Current_Region":i["wd:Region"],
                                    "Employee_Current_Pay_Group":i["wd:Pay_Group"]}
                                    count+=1
                    except :
                        pass
        except:
            print("Error on getting values\n")
    except: 
         return "Error at time to read the XML file // // "
    else:
        Save_Sheet(employee_data,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Data_From_WorkDay", 0, 0)
        return employee_data     
'''

def get_dict_with_condition(employee_data,field,value):
    logger.info("Removing data that doesn't complies with condition")
    try:
        employees_with_benefits = {}
        count = 0
        for i in employee_data.keys():
            if str(employee_data[i][field]).upper() == str(value).upper():
                employees_with_benefits[count]=employee_data[i]
                count+=1
    except:
        logger.info("Error on reading dic")
        return "Error : Cant set the dictionary // // "
    return employees_with_benefits




   
def get_data_before_cutoff(employee_data,cutoffdate):
    logger.info("removing data before cutoff")
    try:
        employees_with_benefits = {}
        count = 0
        for i in employee_data.keys():
            try:
                dateHR =  datetime.strptime(employee_data[i]["Employee_HE_Date"][0:-6], '%Y-%m-%d')
            except:
                dateHR =  datetime.strptime(employee_data[i]["Employee_HE_Date"], '%Y-%m-%d')
            if dateHR <= cutoffdate:
                employees_with_benefits[count]=employee_data[i]
                count+=1
    except:
        logger.info("Error on get cuttof")
        return "Error cant get employees before cutoff // // "
    return employees_with_benefits


'''

def order_dic(dic):
        logger.info("Ordering data")
        print(dic.keys())
    #try:
        list_from_dict = sorted(dic.values(), key=itemgetter("ID", "Date"), reverse=True)
        out_dic = {}
        for i in range(len(list_from_dict)):
            out_dic[i] = list_from_dict[i]
    #except:
    #    return "Error : cant order dictionary // // "
        return out_dic
   
'''

def order_dic(dic,rev):
    logger.info("Ordering data")
    #print(dic.keys())
    try:
        list_from_dict = sorted(dic.values(), key=itemgetter("Employee_ID", "Employee_HE_Date"), reverse=rev)
        out_dic = {}
        for i in range(len(list_from_dict)):
            out_dic[i] = list_from_dict[i]
    except:
        logger.info("Cant order dic")
        return "Error : cant order dictionary // // "
    return out_dic
   
   
def order_pivot_dic(dic):
    logger.info("Ordering data")
    try:
        list_from_dict = sorted(dic.values(), key=itemgetter("ID","Entity", "Dpt"), reverse=False)
        out_dic = {}
        for i in range(len(list_from_dict)):
            out_dic[i] = list_from_dict[i]
    except:
        logger.info("Cant order dic")
        return "Error : cant order dictionary // // "
    return out_dic
     
def order_journal_dic(dic):
        logger.info("Ordering data")
    #try:
        list_from_dict = sorted(dic.values(), key=itemgetter("Entity", "Dpt"), reverse=False)
        out_dic = {}
        for i in range(len(list_from_dict)):
            out_dic[i] = list_from_dict[i]
    #except:
    #    logger.info("Cant order dic")
    #    return "Error : cant order dictionary // // "
        return out_dic

  

def create_pivot_data(Employee_dic,quarters,proration,taxes,cutoff,quarter_to_calculate,company_performances,year):
        logger.info("Creating pre-pivot")
    #try:
        commission=0
        first_id = -1
        Original_Journal = {}
        Converted_Journal = {}
        data = []
        first_data =[]
        count = 0
        ft = 0
        currency_dic={}
        last_date = None
        with open(f"{data_path}\\calculation.txt","w") as f:
            for i in Employee_dic.keys():
                data_to_write=""
                previous_date = int(0)
                first_data = get_employee_data_current_employee(Employee_dic[i])
                employee_current_id = int(first_data[0])
                if  employee_current_id != first_id:
                    data_to_write+=f"+Starting calculations for the employee with id --> {employee_current_id}\n"
                    data = first_data
                    First_calculation = 0
                    first_id = employee_current_id 
                    from_currency = first_data[4]
                    to_currency = first_data[4]
                    dpt= int(first_data[3])
                    Entity = int(first_data[2])
                    #data_to_write+=f"Data: {data}....\n")
                    ft=1
                else:
                    data_to_write+=f"Continuing calculations for the employee with id --> {employee_current_id}\n"
                    phr=data[1]
                    previous_date = datetime.strptime(phr,"%Y-%m-%d-%H:%M") 
                    from_currency = first_data[4]
                    #to_currency = data[4]
                    if from_currency != to_currency:
                        print(f"Employee with id {employee_current_id} have changed from {from_currency} to {to_currency}")
                    dpt = int(data[3])
                    Entity = int(data[2])
                    ft=0
                    #data_to_write+=f"Previous Data:{data} ..... \nCurrent Employee: {first_data}... Previous and current row data Must be different but the same ID\n")
                if from_currency not in currency_dic.keys(): 
                    currency_dic[from_currency ] = {to_currency:0}
                    currency_dic[to_currency ] = {from_currency:0}
                else:
                    currency_dic[from_currency].update({to_currency:0})
                    currency_dic[to_currency].update({from_currency:0})
                start_dates=quarters[0][0]
                end_date = quarters[quarter_to_calculate-1][1]
                hr=first_data[1][:-6]
                hr=datetime.strptime(hr,"%Y-%m-%d")
                #temporal_item = get_partial_amount(quarter_to_calculate, first_data, data, quarters,proration,cutoff,company_performances)
                conversion_rate = get_currency_exchange(from_currency,to_currency,quarters[quarter_to_calculate-1][0].strftime("%Y-%m-%d"))#end_date.strftime("%Y-%m-%d"))
                conversion_rate2 = get_currency_exchange(to_currency,from_currency,quarters[quarter_to_calculate-1][0].strftime("%Y-%m-%d"))
                currency_dic[from_currency][to_currency] = conversion_rate
                currency_dic[to_currency][from_currency] = conversion_rate2
                #data_to_write+=f"Conversion rate {conversion_rate} ::: From currency {from_currency} to currency {to_currency}\n")
                print(first_data[5])
                #time.sleep(50)
                if  "Fixed Percent" in first_data[5]:
                    print("Bonus")
                    data_to_write+=f"Salary for employee {employee_current_id} : {first_data[6]}\n"
                    data_to_write+=f"Percentage of Bonus for employee {employee_current_id} : {first_data[7]}\n"
                    bonus_amount= float(first_data[6]) * float(first_data[7])
                    data_to_write+=f"Initial calculation: Bonus Amount * Percentage [{bonus_amount}]\n"
                    commission=0
                else:
                        print("commission")
                        commission = 1
                        bonus_amount = 0
                if from_currency != to_currency:
                    data_to_write+=f"Conversion from {from_currency} to {to_currency} : {conversion_rate}\n"
                    bonus_amount = bonus_amount * conversion_rate
                    data_to_write+=f"{from_currency} converted to {to_currency} --> {bonus_amount}\n"
                total_amount = 0
                #partial_calculation =  (float(bonus_amount) * company_performances) / 365
                data_to_write+=f"Company Performance: {company_performances}\n"
                partial_calculation =  (float(bonus_amount) * (float(company_performances) / 100) ) / 365
                data_to_write+=f"Partial Calculation = ( Initial Calculation [{bonus_amount}] *  ( Company performance [{company_performances}] / 100 ) ) / 365  ---> {partial_calculation}  :::: Salary per day \n"
                #data_to_write+=f"How much the employee {employee_current_id} must be paid for each day? --> {partial_calculation}\n")
                #with open (f"{data_path}\\calculation.txt","a+") as cal:
                percen_quarter_worked = 0
                id=data[0]
                #previous_id= first_data[0]
                #start_d = quarters[i][0]
                end_d = quarters[quarter_to_calculate-1][1]
                #if hr >= quarters[i][1]:
                #    hr = quarters[i][1]
                if hr <= start_dates:
                    hr = start_dates
                if type(previous_date)==int :
                    #cal.write(f"First time calculation for id {id}\n")
                    #cal.write(f"Start: {start_d} , Hire: {hr} , End: {end_d}\n")
                    if hr <= end_d:
                        percen_quarter_worked = calculate_quarter_proration(hr,quarters[quarter_to_calculate-1][1])
                        if First_calculation == 0:
                            First_calculation = 1
                            percen_quarter_worked += 1
                        #data_to_write+=f"How many days have the employee {employee_current_id} worked for this period? --> {percen_quarter_worked}   ..... Must Be the FIRST CALCULATION\n")
                        data_to_write+=f"Days Worked from {hr} to {quarters[quarter_to_calculate-1][1]} --> {percen_quarter_worked}\n"
                        #cal.write(f"Percent for this quarter: {percen_quarter_worked}\n")
                else:
                    #cal.write(f"No First time calculation for id {id} --> previous id {previous_id}\n")
                    #cal.write(f"Start: {start_d} , Hire: {hr} ,  Previous:{previous_date}, End: {end_d}\n")
                    #already_calculated = calculate_quarter_proration(start_d,previous_date,end_d)
                    #new_calculation = calculate_quarter_proration(start_d,hr,end_d)
                    #print(f"Calculated: {already_calculated}, New Calculation:{new_calculation}\n")
                    if hr  !=  previous_date:
                        if hr <=  quarters[quarter_to_calculate-1][1]:
                            if previous_date >= quarters[quarter_to_calculate-1][1]:
                                previous_date = quarters[quarter_to_calculate-1][1] 
                            if previous_date <= start_dates:
                                previous_date = start_dates
                            percen_quarter_worked =  calculate_quarter_proration(hr,previous_date)
                            if First_calculation == 0:
                                First_calculation = 1
                                percen_quarter_worked += 1
                            #data_to_write+=f"How Many Days have Worked {employee_current_id} --> {percen_quarter_worked}\n")
                            data_to_write+=f"Days Worked from  {hr}  to {previous_date} --> {percen_quarter_worked}\n"
                            #cal.write(f"Calculated: {already_calculated}, New Calculation:{new_calculation}, Real Calculation:{percen_quarter_worked}\n" )
                calculation_per_line = partial_calculation * percen_quarter_worked
                data_to_write+=f"Days [{percen_quarter_worked}]  * Daily Rate [{partial_calculation}]: {calculation_per_line}\n"
                total_amount +=  calculation_per_line
                #data_to_write+=f"How much the employee {employee_current_id} must be paid for this calculation? --> {total_amount}\n")
                #cal.write(f"Subtotal: {total_amount} \t")
                total_amount = round(total_amount)
                data_to_write+=f"Rounded: {total_amount}\n\n\n"
                #data_to_write+=f"Total amount the employee {employee_current_id} must be paid? --> {total_amount}\n\n\n")
                #cal.write(f"\nTotal:{total_amount}\n\n")
                hr=hr.strftime("%Y-%m-%d")
                #if total_amount > 0:
                #if temporal_item != {0}:
                #    Original_Journal[count] = temporal_item 
                #    count +=1
                data=first_data
                if int(percen_quarter_worked) > 0 and commission == 0:
                    f.write(data_to_write)
                    Original_Journal[count] = {"ID" : id,"Date" : hr,'Entity' : Entity ,'Dpt': dpt, 'Currency' : to_currency , 'Amount' : total_amount }  
                    count +=1
        reversed_file=""
        file_string  = ""
        with open(f"{data_path}\\calculation.txt","r") as f:
            file_string = f.read()
        #print(file_string,"\n\n\n","READ")
        file_string = file_string[1:]
        #print(file_string,"\n\n\n")
        file_string = file_string.split("+")[1:]
        #print(file_string,"\n\n\n","READ")
        #counter=1
        #print(len(file_string),"LEN")
        #file_string = file_string[:-1]
        #for i in range(0,len(file_string)):
        #    counter+=1
        #    print(i,"\n\n\n","Counting",len(file_string),counter)
        #    file_string+=str(i)
        #new_lis = new_list.reverse()
        file_string.reverse()
        file_string = "\n".join(x for x in file_string)
        #print(file_string,"\n\n\n")
        #for i in new_list:
        #    reversed_file += i
        #print(reversed_file,"\n\n\n")
        with open(f"{data_path}\\calculation.txt","w") as f:
            f.write(file_string)        
    #except:
    #    logger.info("Error on creating pivot")
    #    year =  int(quarters[0][0].strftime("%Y")) + 1 
    #    Save_Sheet(currency_dic,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Currency_exchange", 0, 0)
    #    return "Error Cant create the pivot data // // "
        year =  int(quarters[0][0].strftime("%Y")) + 1 
        Save_Sheet(currency_dic,f"{ROOT_DIR}\\Data\\Bonus_Report_RaaS_{year}.xlsx",f"Currency_exchange", 0, 0)
        return Original_Journal
       
def get_employee_data_current_employee(Employee):
    try:
        return [Employee['Employee_ID'], Employee["Employee_HE_Date"],  Employee["Employee_Current_Company_ID"],  Employee["Employee_Current_DPT"], Employee["Employee_Current_Currency"], Employee["Employee_Current_Bonus_Plan"], Employee["Employee_Current_Salary"], Employee["Employee_Current_Bonus_Percent"], Employee[ "Employee_Current_Commission"]]
    except:
        logger.info("Error on create data list")
        return "Error: on create employee list"
'''
def get_partial_amount(quarter_to_calculate, first_data, data, quarters,proration,cutoff,company_performances):
    #try:
       
        if total_amount > 0:
            return  {'Entity' : Entity ,'Dpt': dpt, 'Currency' : currency , 'Amount' : total_amount }  
        else:
            return {0}
    #except:
    #    return "Error Cant get partial amount // // "
'''
        
def do_changes():
    pass
'''
#printbonus_amount = bonus_amount * convertion_rate
    if data == first_data:
            dpt= int(first_data[3])
            Entity = int(first_data[2])
            currency = first_data[4]
            hr=first_data[1]
            hr=datetime.strptime(hr,"%Y-%m-%d-%H:%M")
        else:
            dpt = int(data[3])
            Entity = int(data[2])
            currency = data[4]
            hr=first_data[1]
            hr=datetime.strptime(hr,"%Y-%m-%d-%H:%M")
            phr=data[1]
            previous_date = datetime.strptime(phr,"%Y-%m-%d-%H:%M")
    conversion_rate = get_currency_exchange(from_currency,currency,end_date.strftime("%Y-%m-%d"))
'''


'''        
def calculate_quarter_proration(start_date,date_of_hire_rehire,end_date):
    logger.info("Calculating prorration")
    if len(f"{start_date}") > 10 :
        start_date = f"{start_date}"[0:10]
    if len(f"{end_date}") > 10 :
        end_date = f"{end_date}"[0:10]
    if len(f"{date_of_hire_rehire}") > 10 :
        date_of_hire_rehire = f"{date_of_hire_rehire}"[0:10]    
    start_date=datetime.strptime(f"{start_date}","%Y-%m-%d")
    date_of_hire_rehire=datetime.strptime(f"{date_of_hire_rehire}","%Y-%m-%d")
    end_date=datetime.strptime(f"{end_date}","%Y-%m-%d")
    try:
        if date_of_hire_rehire > end_date:
            return 0
        elif date_of_hire_rehire <= start_date:
            return 1
        else:
            diff_hr_start = (date_of_hire_rehire - start_date).days
            diff_start_end = (end_date - start_date).days
            calculation = 1 - ( float (  diff_hr_start / diff_start_end ) )
            return calculation
    except:
        logger.info("Error on calculate prorration")
        return "Error cant calculate quarter proration // // "

'''

        
def calculate_quarter_proration(start_date,end_date):
    logger.info("Calculating prorration")
    try:
        if len(f"{start_date}") > 10 :
            start_date = f"{start_date}"[0:10]
        if len(f"{end_date}") > 10 :
            end_date = f"{end_date}"[0:10]
        start_date=datetime.strptime(f"{start_date}","%Y-%m-%d")
        end_date=datetime.strptime(f"{end_date}","%Y-%m-%d")    
        calculation = (end_date - start_date).days 
        return calculation
    except:
        logger.info("Error on calculate prorration")
        return "Error cant calculate quarter proration // // "









    
def get_currency_exchange(from_currency,to_currency,date_of_conversion):
    logger.info("Getting currency")
    try:
        if from_currency == to_currency:
            return 1
        from_currency = from_currency.strip()
        to_currency = to_currency.strip()
        currentDateTime = datetime.now() - relativedelta(months=1) 
        if type(date_of_conversion) == str: 
            date_of_conversion = datetime.strptime(date_of_conversion, '%Y-%m-%d')
        if date_of_conversion >= currentDateTime:
            date_of_conversion = currentDateTime.strftime('%Y-%m-%d')
        else:
            date_of_conversion = date_of_conversion.strftime('%Y-%m-%d')
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/currencyRates?finder=CurrencyRatesFinder;fromCurrency={from_currency},toCurrency={to_currency},userConversionType=Daily Rate,startDate={date_of_conversion},endDate={date_of_conversion}&onlyData=True"
        #print(url)
        headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
        user,passw  = read_credentials()
        #print(user,passw)
        ret = requests.get(url, auth=(user, passw), headers=headerss)
        #print(ret)
        #print(ret.text)
        rate = float(json.loads(ret.text)["items"][0]["ConversionRate"])
    except:
        logger.info("Error cant rate conversion")
        return "Error : cant calculate rate conversion // "
    return rate


def create_pivot(pivot_data):
    logger.info("CReating pivot")
    logger.info("Creating Pivot")
    pivot={}
    try:
        for i in pivot_data.keys():
            entity = int(pivot_data[i]["Entity"])
            dpt= int(pivot_data[i]["Dpt"])
            currency = pivot_data[i]["Currency"]
            amount= Decimal(pivot_data[i]["Amount"])
            if entity not in pivot.keys():
                pivot[entity] =  {  dpt: { 'Currency': currency, 'Amount' : Decimal(amount) } }
            else:
                if dpt in pivot[entity].keys():
                    total_amount = Decimal(pivot[entity][dpt]["Amount"]) + amount
                    pivot[entity][dpt]= { 'Currency': currency, 'Amount' : Decimal(total_amount)}  
                else:
                    pivot[entity].update({ dpt : { 'Currency': currency, 'Amount' : amount } })  
    except:
        logger.info("Cant create pivot")
        return "Error : can create the pivot, maybe one of the parameters has not been set or its wrong (e.g cutoff date). // // "    
    return pivot
    

def create_Journals(pivot,end_date,reversal,quarter_to_calculate,taxes,quarters):
        logger.info("creating Journal")
    #try:
        entities=get_entities("primary")
        jrnl={}
        tx_jrn={}
        usd_jrnl={}
        tx_usd_jrnl={}
        count=0
        Year_To_calculate = int(end_date.strftime("%Y"))
        year = f"{Year_To_calculate}"[:-2]
        Base_Bach_Name= f"Q{quarter_to_calculate}'{Year_To_calculate} - Bonus Accrual"
        Batch_Name = Base_Bach_Name
        period = end_date.strftime("%b-%y")
        state= Get_Ledger_Status(f"{period}")
        if "Error" in state:
            print("Error on state")
            return [f"{state} // "]
        for i in pivot.keys():
            data = entities[f"{i}"]
            #print(f"data : {data}")
            group_id= datetime.today().strftime('%Y%m%d%H%M%S')
            Ledger_Country = data["Name"].split(" ")[1]
            Journal_Name = f"{Base_Bach_Name} - {Ledger_Country}"
            Ledger_Description =  f"{Base_Bach_Name} - {Ledger_Country} {group_id}"
            Ledger_tax_Description =  f"{Base_Bach_Name} - Tax - {Ledger_Country} {group_id}"
            Ledger_converted_Description =  f"{Base_Bach_Name} - Adjustment - {Ledger_Country} {group_id}"
            Ledger_tax_converted_Description =  f"{Base_Bach_Name} - Tax Adjustment -{Ledger_Country} {group_id}"
            Ledger_name = data["Name"]
            Ledger_id=data["Ledger"]
            Accounting_date = end_date.strftime("%Y-%m-%d")
            #datetime.datetime.now().strftime("%m/%d/%Y")
            Source = "AA Bot"
            Category  = "COMPandBEN_Accrual"
            entity = int(i)
            entity_total= Decimal(0.0)
            total_tax= Decimal(0.0)
            converted_total= Decimal(0.0)
            tax_converted_total= Decimal(0.0)
            header_soap = Create_SOAP_HEADER(Batch_Name,Ledger_id,period,Accounting_date)
            header_soap_tx = Create_SOAP_HEADER(f"{Batch_Name}_Tax",Ledger_id,period,Accounting_date)
            header_soap_us = Create_SOAP_HEADER(f"{Batch_Name}-Adjustment",Ledger_id,period,Accounting_date)
            header_soap_us_tx = Create_SOAP_HEADER(f"{Batch_Name} Tax-Adjustment",Ledger_id,period,Accounting_date)
            journal_line=""
            journal_tax_line=""
            journal_converted_line =""
            journal_tax_converted_tax_line =""
            dpt_lst=[]
            #print("\n")
            #print(i)
            deb = []
            tx=[]
            deb_conv=[]
            tx_conv=[]
            for j in pivot[i].keys():
                dpt=j
                dpt_lst.append(int(j))
                account="60200"
                tx_account="60810"
                prod="000000"
                proj="000000"
                ico="00"
                fut="0"
                currency= pivot[i][j]["Currency"]
                #debit= float(pivot[i][j]["Amount"])
                debit= round(Decimal(pivot[i][j]["Amount"]),2)
                #debit = set_n_digits(debit)
                tax =  round(debit * taxes[i][quarter_to_calculate-1],2)
                #tax = set_n_digits(tax)
                #print(tax)
                CR=get_currency_exchange(currency,"USD",quarters[quarter_to_calculate-1][0])
                print(CR)
                conversion_rate = round(CR)# end_date)   )
                converted = debit * conversion_rate
                #converted = set_n_digits(converted)
                tax_converted = tax * conversion_rate
                #tax_converted = set_n_digits(tax_converted)
                if "JP" in Ledger_name:
                    debit=round(debit)
                    tax=round(tax)
                    converted=round(converted)
                    tax_converted=round(tax_converted)
                    tx_account="60870"
                deb.append(debit)
                entity_total += debit
                total_tax += tax
                tx.append(tax)
                #total_tax = entity_total * taxes[i][quarter_to_calculate-1]
                converted_total += converted
                deb_conv.append(converted)
                tax_converted_total += tax_converted
                tx_conv.append(tax_converted)
                #tax_converted_total = round(total_tax * round(Decimal(conversion_rate),3),2)
                '''
                entity_total.append(debit)
                total_tax.append(tax)
                converted_total.append(converted)
                tax_converted_total.append(tax_converted
                debit = round(debit,2)
                tax =  round(debit,2) * round(float(taxes[i][quarter_to_calculate-1]),2)
                tax = round(tax,2)
                #print(tax)
                conversion_rate = round(  get_currency_exchange(currency,"USD",end_date) ,2  )
                converted = round(debit,2) * round(conversion_rate,2)
                converted = round(converted,2)
                tax_converted = round(tax,2) * round(conversion_rate,2)
                tax_converted = round(tax_converted,2)
                if "JP" in Ledger_name:
                    debit=round(debit)
                    tax=round(tax)
                    converted=round(converted)
                    tax_converted=round(tax_converted)
                entity_total.append(debit)
                total_tax.append(tax)
                converted_total.append(converted)
                tax_converted_total.append(tax_converted)
                '''
                
                #print(f"Entity:{entity}, Dpt: {dpt}, Debit: {debit}, Tx: {tax}, Convertion: {converted}, Tax Converted:{tax_converted}, Conversion:{convertion_rate}")
                #print(f"Entity Total: {entity_total}, Total Tax: {tax}, Converted Total: {converted_total} , Tx Converted Total: {tax_converted_total}")
                
                #print(tax)
                #print(f"\t\t{tax_converted}")
                journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, debit, reversal, account, entity, dpt, Journal_Name,Ledger_name,Ledger_Description,0)
                journal_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, tax, reversal, tx_account, entity, dpt, f"{Journal_Name}_Tax",Ledger_name,Ledger_tax_Description,0)
                journal_converted_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", converted, reversal, account, entity, dpt, f"{Journal_Name} - Adjustment",Ledger_name,Ledger_converted_Description,0)
                journal_tax_converted_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", tax_converted, reversal, tx_account, entity, dpt, f"{Journal_Name} Tax- Adjustment",Ledger_name,Ledger_tax_converted_Description,0)

                jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : debit,"Credit":"0", "Reversal" : reversal,"Account" : account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : Journal_Name, "Journal_Name" : Ledger_name,"Journal_description":Ledger_Description }
                tx_jrn[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : tax,"Credit":"0", "Reversal" : reversal,"Account" : tx_account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name}_Tax", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_Description}
                usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : converted,"Credit":"0", "Reversal" : reversal,"Account" : account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name} - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_converted_Description}
                tx_usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : tax_converted,"Credit":"0", "Reversal" : reversal,"Account" : tx_account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name} Tax - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_converted_Description}
                count += 1
            #print(deb)
            #print(sum(deb))
            print(entity_total)
            #print(tx)
            #print(sum(tx))
            print(total_tax)
            #print(deb_conv)
            #print(sum(deb_conv))
            print(converted_total)
            #print(tx_conv)
            #print(sum(tx_conv))
            print(tax_converted_total)
            dpt = "000"
            account="21030"
            tx_account="21110"
            #entity_total_s = sum(entity_total)
            #total_tax_s = sum(total_tax)
            #converted_total_s=sum(converted_total)
            #tax_converted_total_s = sum(tax_converted_total)
            
            #test=0
            #for t in  total_tax:
            #    test += t
            #    #print(t,test)
            #print(test,total_tax_s,sum(total_tax))
            '''
            print(f"entity:{entity}, Dpt:{dpt}, Entity Total: {entity_total_s}, Total Taxes: { total_tax_s}, Total Converted : {converted_total_s}, Total Tax Converted: {tax_converted_total_s}, Conversion:{conversion_rate}")
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, entity_total_s, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
            journal_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, 0 total_tax_s, reversal, tx_account, entity, dpt, f"{Journal_Name}_Tax",Ledger_name,Ledger_tax_Description,1)
            journal_converted_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", converted_total_s, reversal, account, entity, dpt, f"{Journal_Name} - Adjustment",Ledger_name,Ledger_converted_Description,1)
            journal_tax_converted_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", tax_converted_total_s, reversal, tx_account, entity, dpt, f"{Journal_Name} Tax- Adjustment",Ledger_name,Ledger_tax_converted_Description,1)
            0
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":entity_total_s, "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : Journal_Name, "Ledger_Name" : Ledger_name,"Journal_description": Ledger_Description}
            tx_jrn[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit": total_tax_s, "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name}_Tax", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_Description}
            usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":converted_total_s, "Reversal" : reversal,"Account" : account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_converted_Description}
            tx_usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":tax_converted_total_s, "Reversal" : reversal,"Account" : tx_account, "Entity" :entity, "Dpt" : dpt, "Jrnl_Name" : f"{Journal_Name} Tax - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_converted_Description}
            count += 1
            '''
            #print(f"entity:{entity}, Dpt:{dpt}, Entity Total: {entity_total}, Total Taxes: { total_tax}, Total Converted : {converted_total}, Total Tax Converted: {tax_converted_total}, Conversion:{convertion_rate}")
            journal_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency, entity_total, reversal, account, entity, dpt,Journal_Name, Ledger_name,Ledger_Description,1)
            journal_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, currency,  total_tax, reversal, tx_account, entity, dpt, f"{Journal_Name}_Tax",Ledger_name,Ledger_tax_Description,1)
            journal_converted_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", converted_total, reversal, account, entity, dpt, f"{Journal_Name} - Adjustment",Ledger_name,Ledger_converted_Description,1)
            journal_tax_converted_tax_line+=Create_SOAP_Body(Ledger_id, period, Accounting_date, group_id, "USD", tax_converted_total, reversal, tx_account, entity, dpt, f"{Journal_Name} Tax- Adjustment",Ledger_name,Ledger_tax_converted_Description,1)
      
            jrnl[count]={"Batch_Name" : Base_Bach_Name, "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit":entity_total, "Reversal" : reversal,"Account" : account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : Journal_Name, "Ledger_Name" : Ledger_name,"Journal_description": Ledger_Description}
            tx_jrn[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : currency, "Debit" : "0","Credit": total_tax, "Reversal" : reversal,"Account" : tx_account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name}_Tax", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_Description}
            usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":converted_total, "Reversal" : reversal,"Account" : account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name} - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_converted_Description}
            tx_usd_jrnl[count]={"Batch_Name" : f"{Base_Bach_Name}", "Ledger_Id" : Ledger_id, "Period" : period, "Accounting_Dte" : Accounting_date, "Group_Id" : group_id, "Currency" : "USD", "Debit" : "0","Credit":tax_converted_total, "Reversal" : reversal,"Account" : tx_account, "Entity" :f"{entity}", "Dpt" : f"{dpt}", "Jrnl_Name" : f"{Journal_Name} Tax - Adjustment", "Ledger_Name" : Ledger_name,"Journal_description": Ledger_tax_converted_Description}
            count += 1
            
            #print(total_tax)
            #print(f"\t\t{tax_converted_total}")
            foot= Create_SOAP_Foot()
            
            journal_entry = f"{header_soap}{journal_line}{foot}"
            journal_tx_entry = f"{header_soap_tx}{journal_tax_line}{foot}"
            ####journal_converted_entry = f"{header_soap_us}{journal_converted_line}{foot}"
            ####journal_tx_converted_entry = f"{header_soap_us_tx}{journal_tax_converted_tax_line}{foot}"
            #with open(f"{data_path}\\journal.txt","a+") as jrn:
            #    jrn.write(journal_tx_entry)
            logger.info("Journal lines pushed")
            response_to_push = Push_To_Oracle(journal_entry)
            if ">0</result>" not in response_to_push.text:
                #print(response_to_push.text)
                print("Error on pushing")
                return ["Error on pushed Journal lines //"]
            logger.info("importing journal lines")
            response_to_push = Push_To_Oracle(journal_tx_entry)
            if ">0</result>" not in response_to_push.text:
                #print(response_to_push.text)
                print("Error on pushing")
                return ["Error on pushed Journal lines //"]
            ####response_to_push = Push_To_Oracle(journal_converted_entry)
            ####if ">0</result>" not in response_to_push.text:
            ####    #print(response_to_push.text)
            ####    #print("Error on pushing")
            ####    return ["Error on pushed Journal lines //  "]
            ####response_to_push = Push_To_Oracle(journal_tx_converted_entry)
            ####if ">0</result>" not in response_to_push.text:
            ####    #print(response_to_push.text)
            ####    print("Error on pushing")
            ####    return ["Error on pushed Journal lines // "]
            importing = Create_Import_SOAP(set_id, source_id,Ledger_id , f"{group_id}")
            response_import = Import_to_GL(importing)
            import_id = json.loads(response_import.text)["ReqstId"]
            #print(import_id)
            print(reversal)
            if import_id != -1:
                while True:
                    Status = json.loads(Get_Status(import_id).text)["items"][0]["RequestStatus"]
                    #print(Status)
                    if Status == "SUCCEEDED":
                        #print("Imported")
                        logger.info("journal lines imported")
                        break
                    if Status == "ERROR" or Status == "WARNING":
                        #print("Error on import")
                        return [f"Oracle's Error: Warning or Error on Importing PID : {import_id}   Ledger ID: {Ledger_id} on journal // "]
            else:
                logger.info("Error on import Journal lines")
                print("Oracle's Error on import")
                return "Error on import Journal lines // //"
        return [jrnl,tx_jrn,usd_jrnl,tx_usd_jrnl]
    #except:
    #    return ["Error cant create journal // "]
        
    
def Create_SOAP_HEADER(Batch_Name, ledger_id, period_name, accounting_date):
    logger.info("Creating SOAP HEADER")
    return f'''<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:typ="http://xmlns.oracle.com/apps/financials/generalLedger/journals/desktopEntry/journalImportService/types/" xmlns:jour="http://xmlns.oracle.com/apps/financials/generalLedger/journals/desktopEntry/journalImportService/" xmlns:jour1="http://xmlns.oracle.com/apps/flex/financials/generalLedger/journals/desktopEntry/journalLineGdf/">
   <soapenv:Header/>
   <soapenv:Body>
      <typ:importJournals>
         <typ:interfaceRows>
            <jour:BatchName>{Batch_Name}</jour:BatchName>
            <jour:BatchDescription>{Batch_Name}</jour:BatchDescription>
            <jour:LedgerId>{ledger_id}</jour:LedgerId>
            <jour:AccountingPeriodName>{period_name}</jour:AccountingPeriodName>
            <jour:AccountingDate>{accounting_date}</jour:AccountingDate>
            <jour:UserSourceName>AA Bot</jour:UserSourceName>
            <jour:UserCategoryName>COMPandBEN_Accrual</jour:UserCategoryName>
            <jour:ErrorToSuspenseFlag>True</jour:ErrorToSuspenseFlag>
            <jour:SummaryFlag>True</jour:SummaryFlag>
            <jour:ImportDescriptiveFlexField>N</jour:ImportDescriptiveFlexField>
'''


def Create_SOAP_Body(ledger_id, period_name, accounting_date, group_id, currency, amount, reversal, account, ent, dep, Journal_name,ledger_name,Ledger_Description,type):
    logger.info("Creating SOAP Body")
    if type==0:
        crdr = "Dr"
    if type == 1:
        crdr = "Cr"
    return f'''               <jour:GlInterface>
               <jour:LedgerId>{ledger_id}</jour:LedgerId>
               <jour:LedgerName>{ledger_name}</jour:LedgerName>
               <jour:PeriodName>{period_name}</jour:PeriodName>
               <jour:AccountingDate>{accounting_date}</jour:AccountingDate>
               <jour:UserJeSourceName>AA BOT</jour:UserJeSourceName>
               <jour:UserJeCategoryName>COMPandBEN_Accrual</jour:UserJeCategoryName>
               <jour:GroupId>{group_id}</jour:GroupId>
               <jour:ChartOfAccountsId/>
               <jour:BalanceType>A</jour:BalanceType>
               <jour:CodeCombinationId/>
               <jour:Segment1>{ent}</jour:Segment1>
               <jour:Segment2>{dep}</jour:Segment2>
               <jour:Segment3>{account}</jour:Segment3>
               <jour:Segment4>000000</jour:Segment4>
               <jour:Segment5>000000</jour:Segment5>
               <jour:Segment6>00</jour:Segment6>
               <jour:Segment7>0</jour:Segment7>
               <jour:CurrencyCode>{currency}</jour:CurrencyCode>
               <jour:Entered{crdr}Amount currencyCode="{currency}">{amount}</jour:Entered{crdr}Amount>
               <jour:AccountedCr/>
               <jour:AccountedDr/>
               <jour:UserCurrencyConversionType>User</jour:UserCurrencyConversionType>
               <jour:CurrencyConversionDate>{accounting_date}</jour:CurrencyConversionDate>
               <jour:CurrencyConversionRate>1</jour:CurrencyConversionRate>	
               <jour:Reference4>{Ledger_Description}</jour:Reference4> 
               <jour:Reference5>{Ledger_Description}</jour:Reference5>
               <jour:Reference7>Y</jour:Reference7>          
               <jour:Reference8>{reversal}</jour:Reference8>       
               <jour:Reference9>Y</jour:Reference9>
            </jour:GlInterface>\n'''
            
def Create_SOAP_Foot():
    logger.info("Creating SOAP foot")
    return '''         </typ:interfaceRows>
      </typ:importJournals>
   </soapenv:Body>
</soapenv:Envelope>\n
'''





def Push_To_Oracle(soap):
    logger.info("Pushing to oracle")
    try:
        url = f"{oracle_url}/fscmService/JournalImportService?WSDL"
        headerss = {"Content-Type": "text/xml;charset=UTF-8",
                    "Accept-Encoding": "gzip, deflate",
                    "Connection": "Keep-Alive"
                    }
        logger.info("Read credentials")
        user,passw  = read_credentials()
        return requests.post(url, data=soap, auth=(user, passw), headers=headerss)
    except:
        return f"Error on push oracle"



def AutoPost(soap):
    try:
        logger.info("Trying to post")
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
        headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
        user,passw  = read_credentials()
        return requests.post(url, data=soap, auth=(user, passw), headers=headerss)
    except:
        return "Error on autopost"


def Get_Status(RqstId):
    logger.info("Trying to get the PID status")
    try:
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations?finder=ESSJobStatusRF;requestId={RqstId}"
        headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
        user,passw  = read_credentials()
        return requests.get(url, auth=(user, passw), headers=headerss)
    except:
        return f"Error on get status"


def Create_AutoPost():
    logger.info("Creating Autopost payload")
    return f'''{{
    "OperationName":"submitESSJobRequest",
    "JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common/",
    "JobDefName":"AutomaticPosting",
    "ESSParameters":"{Autopost_id}",
    "ReqstId":null
    }} '''
    
def Create_Import_SOAP(set_id, source_id, ledger_id, group_id):
    logger.info("Creating import payload")
    return f'''{{ "OperationName":"submitESSJobRequest",
    "JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common/",
    "JobDefName":"JournalImportLauncher",
    "ESSParameters":"{set_id},{source_id},{ledger_id},{group_id},N,N,N",
    "ReqstId":null
    }}
    '''
    
def Import_to_GL(soap):
    logger.info("Importing data")
    try:
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
        headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
        user,passw  = read_credentials()
        return requests.post(url, data=soap, auth=(user, passw), headers=headerss)
    except:
        return f"Error on import to GL"




def Get_Ledger_Status(Period):
    logger.info("Getting ledger status")
    entities = get_entities("primary")
    flg=0
    String = f"Error : Next Ledgers are closed for period {Period} : <br>"
    try:
        for i in entities.keys():
            ledger_status = verify_closed_periods(Period,entities[i]["Ledger"])
            ledger_json=json.loads(ledger_status.text)
            ledger_json = ledger_json["items"][0]
            if ledger_json["ClosingStatus"] == "C":
                flg=1
                closed = entities[i]["Name"]
                String += f"{closed}<br>"
        if flg ==0:
            return "0"
        else:
            logger.info(f"Error : {String}")
            return f" Error : {String} // "
    except:
        return "Error : Not Well formed data // "
    
    
    
def verify_closed_periods(Period,Ledger_id):
    logger.info("Verifing closed ledgers")
    try:
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/accountingPeriodStatusLOV?q=PeriodNameId={Period};ApplicationId=101;LedgerId={Ledger_id}"
        #print(url)
        user,passw  = read_credentials()
        headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }    
        return requests.get(url, auth=(user, passw), headers=headerss)
    except:
        return f"Error on verify closed periods"



def verify(data):
    logger.info(f"Trying to verify data {data}")
    data=data.upper()
    try:
        if "OVERRIDE" in data:
            return "ok"
        else:
            try:
                logger.info("Reading Parameters")
                XLS = read_xls_parameters()
                if type(XLS) == str:
                        return f"Error : Parameter file cant be read // //"
            except:
                    return "Error : Parameter file cant be read"
            else:      
                    start_date = XLS.iloc[0,1]
                    quarter_to_calculate = int(XLS.iloc[0,5])
                    Year_To_calculate = int(start_date.strftime("%Y")) + 1
                    data =  f"{Year_To_calculate}-{quarter_to_calculate}"
                    try:
                        with open(f"{config_path}\\config.cfg", "r") as config:
                            lst = config.read().split("\n")
                            if data in lst:
                                return "Error: This Year-Quarter has been already running and override has not been defined // //"
                    except:
                        with open(f"{config_path}\\config.cfg", "a+") as config:
                            config.write(f"{data}\n")
                            return "Ok"
                    
    except:
        logger.info(f"Data received not well defined")
        logger.info(f"Trying to push the log into the s3_bucket")
        Push_To_S3(f"{logs_path}\\functions.log", "Process4", "Log")
        return "Error : Data Malformed. // "



def Push_To_S3(File, process, subdir):
    '''
    This function push a file into the S3 bucket
    '''
    logger.info("Pushing data to s3")
    try:
        with open(File, "rb") as f:
            key = File.split("\\")[-1]
            response = s3_bucket.upload_fileobj(f, bucket_name, f"{process}/{subdir}/{key}")
            f.close()
    except Exception as e:
        logger.info(f"Error occurs while uploading")
        return "Error : Error occurs while uploading <BR> //"
    
    
def clear():
    try:
        logger.info("Clearing directories")
        files_to_delete = os.listdir(data_path)
        for item in files_to_delete:
            if item.endswith(".xlsx"):
                os.remove(os.path.join(data_path, item))
        for item in files_to_delete:
            if item.endswith(".csv"):
                os.remove(os.path.join(data_path, item))
        for item in files_to_delete:
            if item.endswith(".xml"):
                os.remove(os.path.join(data_path, item))
        for item in files_to_delete:
            if item.endswith(".txt"):
                os.remove(os.path.join(data_path, item))
    except:
        return f"Error on clear the data directory"
            
            
            
def set_two_digits(number):
    #print(number)
    string_number = f"{number}"
    string_number = string_number.split(".")
    Integer = string_number[0]
    Decimal = string_number[1] + "00" 
    Decimal = Decimal[:2] 
    Decimal = int(Decimal)
    Decimal = Decimal/100
    Integer = int(Integer) 
    float_number = Integer + Decimal
    #print(f"{float_number}")
    return float_number

def set_n_digits(v, ndigits=set_ndigits, rt_str=False):
    try:
        d = Decimal(v)
        #v_str = ("{0:.%sf}" % ndigits).format(round(d, ndigits))
        #print(v_str)
        #if rt_str:
        #   return v_str
        #return Decimal(v_str)
        return round(d,ndigits)
    except:
        return f"Error on set the digits"





def Get_Validation(entity, dpt,account, period_name):
    try:
        entities = get_entities("primary");
        name = entities[f"{entity}"]["Name"]
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/ledgerBalances?limit=100000&finder=AccountBalanceFinder;accountCombination={entity}.{dpt}.{account}.000000.000000.00.0,accountingPeriod={period_name},currency=STAT,ledgerSetName={name},mode=Detail&fields=LedgerName,PeriodName,Currency,DetailAccountCombination,Scenario,BeginningBalance,PeriodActivity,EndingBalance,AmountType,CurrencyType,ErrorDetail"
        headerss = {"Content-Type": "application/json",
                    "Connection": "Keep-Alive"
                    }
        user,passw  = read_credentials()
        #print(user,passw)
        return requests.get(url, auth=(user, passw), headers=headerss)
    except:
        return f"Error on get validation"




def transform(year):
    try:
        xmlPath=f"{ROOT_DIR}\\Data\\{year}_bonusfile.xml"
        xslPath = f"{data_path}\\nt1.xsl"
        xslRoot = etree.fromstring(open(xslPath).read())
        transform = etree.XSLT(xslRoot)
        xmldata=open(xmlPath,"rb").read()
        #print(xmldata)
        xmlRoot = etree.fromstring(xmldata)
        transRoot = transform(xmlRoot)
        #print("Almost there")
        with open(f"{data_path}\\newxml.xml","wb") as newxml:
            newxml.write(etree.tostring(transRoot))
    except:
        return f"Error on set the transform"